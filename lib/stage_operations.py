"""Contexto operativo para la UI de Operación (prerequisitos, KPIs, artefactos)."""
from __future__ import annotations

import glob
import os
from datetime import datetime
from typing import Any

import config
import display_format as fmt
import ops_execution_history
import stage_commands
import inbox_gaps
from settings import get_bool_setting

def _month_dir(year: int | str, month: str) -> str:
    return os.path.join(config.RAIZ, str(year), month)


def _es_glosa_provisionado(glosa: object) -> bool:
    text = str(glosa or "").lower()
    return any(token in text for token in ("provisionado", "provisonado", "provs"))


def _excel_row_str(row_obj: Any, df_columns: Any, *names: str) -> str:
    for name in names:
        if name in df_columns:
            return _cell_str(row_obj.get(name))
    return ""


def _glosa_xml_viva(row_obj: Any, month_path: str) -> str:
    """Glosa del XML en disco si hay archivo; si no, columna del Excel."""
    from stages.stage3.revision_core import extraer_datos_xml, glosa_xml_de_fila

    archivo = _cell_str(row_obj.get("archivo_xml") if hasattr(row_obj, "get") else "")
    if archivo and month_path:
        path = os.path.join(month_path, archivo)
        if os.path.isfile(path):
            datos = extraer_datos_xml(path)
            if "error" not in datos:
                glosa = str(datos.get("descripcionLinea", "") or "").strip()
                if glosa:
                    return glosa
    return glosa_xml_de_fila(row_obj, month_path)


def _recepcion_efectivo_fila(
    row_obj: Any, df_columns: Any, month_path: str
) -> tuple[str, str, bool, str, str]:
    """(estado Excel, estado efectivo, glosa coincide, glosa XML usada, modo coincidencia)."""
    from stages.stage3.revision_core import clasificar_coincidencia_glosa

    glosa_pedida = _excel_row_str(row_obj, df_columns, "GLOSA")
    glosa_xml = _glosa_xml_viva(row_obj, month_path)
    glosa_match_mode = "exacta" if not glosa_xml else clasificar_coincidencia_glosa(glosa_pedida, glosa_xml)
    glosa_xml_coincide = glosa_match_mode != "distinta"
    estado_rx = _excel_row_str(row_obj, df_columns, "Estado_Recepcion")
    if estado_rx.upper() == "RECIBIDO" and glosa_xml and not glosa_xml_coincide:
        estado_efectivo = "RECIBIDO CON ERROR"
    else:
        estado_efectivo = estado_rx
    return estado_rx, estado_efectivo, glosa_xml_coincide, glosa_xml, glosa_match_mode


def _inc_recepcion_kpi(recepcion: dict[str, int], estado_efectivo: str) -> None:
    estado = estado_efectivo.upper()
    if estado == "RECIBIDO":
        recepcion["recibido"] += 1
    elif estado == "RECIBIDO CON ERROR":
        recepcion["recibido_con_error"] += 1
    elif estado == "NO RECIBIDO":
        recepcion["no_recibido"] += 1
    elif not estado:
        recepcion["pendiente"] += 1
    else:
        recepcion["otro"] += 1


def _excel_avance_from_db(year: int | str, month: str, *, row_limit: int = 500) -> dict[str, Any] | None:
    """Vista avance desde DB canónica. Retorna None si el período no existe en DB."""
    from sqlalchemy import select
    from sqlalchemy.exc import SQLAlchemyError

    from db.models import Boleta, BoletaXmlData, Docente, Institucion, Periodo
    from db.session import SessionLocal

    month_path = _month_dir(year, month)
    out: dict[str, Any] = {
        "year": int(year) if str(year).isdigit() else year,
        "month": month,
        "month_dir": month_path,
        "solicitud_path": os.path.join(month_path, "Solicitud.xlsx"),
        "solicitud_exists": os.path.isfile(os.path.join(month_path, "Solicitud.xlsx")),
        "sheets": [],
        "solicitud_sheet": None,
        "total_rows": 0,
        "recepcion": {"recibido": 0, "recibido_con_error": 0, "no_recibido": 0, "pendiente": 0, "otro": 0},
        "correo_solicitud": {"enviado": 0, "omitido": 0, "error": 0, "pendiente": 0, "otro": 0},
        "recordatorios": {"con_recordatorio": 0, "total_envios": 0},
        "xml_extract": {"ok": 0, "observacion": 0, "pendiente": 0, "con_archivo": 0},
        "archivos_mes": {"xml": 0, "pdf": 0},
        "pagos": {
            "sheet_exists": False,
            "total_rows": 0,
            "enviado": 0,
            "pendiente": 0,
            "error": 0,
            "omitido": 0,
            "otro": 0,
        },
        "rows": [],
        "rows_truncated": False,
        "read_error": None,
        "mtime": None,
    }
    if os.path.isdir(month_path):
        out["archivos_mes"]["xml"] = len([f for f in os.listdir(month_path) if f.lower().endswith(".xml")])
        out["archivos_mes"]["pdf"] = len([f for f in os.listdir(month_path) if f.lower().endswith(".pdf")])

    try:
        with SessionLocal() as session:
            periodo = session.execute(
                select(Periodo).where(Periodo.anio == int(year), Periodo.mes_nombre == month)
            ).scalar_one_or_none()
            if periodo is None:
                return None
            rows = session.execute(
                select(Boleta, BoletaXmlData, Docente, Institucion)
                .outerjoin(BoletaXmlData, BoletaXmlData.boleta_id == Boleta.id)
                .outerjoin(Docente, Docente.id == Boleta.docente_id)
                .outerjoin(Institucion, Institucion.id == Boleta.institucion_id)
                .where(Boleta.periodo_id == periodo.id)
                .order_by(Boleta.id.asc())
            ).all()
    except SQLAlchemyError:
        return None

    out["total_rows"] = len(rows)
    limit = min(int(row_limit), len(rows)) if row_limit > 0 else len(rows)
    out["rows_truncated"] = len(rows) > limit
    for idx, (bo, xml, docente, institucion) in enumerate(rows[:limit]):
        recep = str(getattr(bo, "recepcion_status", "") or "").upper()
        if recep == "RECIBIDO_OK":
            out["recepcion"]["recibido"] += 1
            estado_ef = "RECIBIDO"
        elif recep == "RECIBIDO_ERROR":
            out["recepcion"]["recibido_con_error"] += 1
            estado_ef = "RECIBIDO CON ERROR"
        elif recep == "NO_RECIBIDO":
            out["recepcion"]["no_recibido"] += 1
            estado_ef = "NO RECIBIDO"
        else:
            out["recepcion"]["pendiente"] += 1
            estado_ef = ""

        xml_status = str(getattr(bo, "xml_status", "") or "").upper()
        if xml_status == "OK":
            out["xml_extract"]["ok"] += 1
            out["xml_extract"]["con_archivo"] += 1
            xml_clase = "ok"
            obs_xml = "Datos extraídos OK"
        elif xml_status == "ERROR":
            out["xml_extract"]["observacion"] += 1
            out["xml_extract"]["con_archivo"] += 1
            xml_clase = "observacion"
            obs_xml = str(getattr(xml, "observaciones_xml", "") or "")
        else:
            out["xml_extract"]["pendiente"] += 1
            xml_clase = "pendiente"
            obs_xml = ""

        has_xml_evidence = bool(
            xml
            and (
                _cell_str(getattr(xml, "numero_boleta", None))
                or _cell_str(getattr(xml, "descripcion_linea", None))
                or _cell_str(getattr(xml, "total_honorarios", None))
                or _cell_str(getattr(bo, "descripcion", None))
            )
        )

        mail_rx = str(getattr(bo, "mail_recepcion_status", "") or "").upper()
        correo_rx = ""
        correo_clase = "pendiente"
        if mail_rx == "ENVIADO_OK":
            correo_rx = "Enviado (confirmación)"
            correo_clase = "enviado"
            out["correo_solicitud"]["enviado"] += 1
        elif mail_rx == "ENVIADO_PROBLEMA":
            correo_rx = "Enviado (observación/reenvío)"
            correo_clase = "error"
            out["correo_solicitud"]["error"] += 1
        else:
            out["correo_solicitud"]["pendiente"] += 1

        sr = getattr(bo, "solicitud_row", None) if isinstance(getattr(bo, "solicitud_row", None), dict) else {}
        name = _solicitud_field(sr, "NAME") or str(getattr(docente, "nombre_completo", "") or "")
        sede = _solicitud_field(sr, "SEDE") or str(getattr(docente, "sede", "") or "")
        email = _solicitud_field(sr, "Email_Docente", "Correo_Personal") or str(
            getattr(docente, "email_personal", "") or ""
        )
        email_dp = _solicitud_field(sr, "Email_DP") or str(getattr(docente, "email_dp", "") or "")
        location = _solicitud_field(sr, "LOCATION")
        nombre_razon = _solicitud_field(sr, "NOMBRE RAZON") or str(getattr(institucion, "nombre_razon", "") or "")
        direccion_razon = _solicitud_field(sr, "DireccionRazon") or str(
            getattr(institucion, "direccion_razon", "") or ""
        )

        row_payload = {
            "row": idx + 2,
            "emplid": fmt.format_rut_cl(getattr(bo, "emplid", "")) or str(getattr(bo, "emplid", "") or ""),
            "rut_sin_dv": fmt.format_rut_sin_dv(getattr(bo, "emplid", "")) or "",
            "name": name,
            "sede": sede,
            "location": location,
            "email": email,
            "email_dp": email_dp,
            "rut_razon": fmt.format_rut_cl(getattr(bo, "rut_razon", "")) or str(getattr(bo, "rut_razon", "") or ""),
            "nombre_razon": nombre_razon,
            "direccion_razon": direccion_razon,
            "glosa": str(getattr(bo, "glosa", "") or ""),
            "provisionado": _es_glosa_provisionado(getattr(bo, "glosa", "")),
            "estado_recepcion": str(getattr(bo, "estado_recepcion", "") or ""),
            "estado_recepcion_efectivo": estado_ef,
            "glosa_xml_coincide": (
                (str(getattr(bo, "glosa_match_mode", "") or "") != "distinta") if has_xml_evidence else None
            ),
            "glosa_match_mode": (
                str(getattr(bo, "glosa_match_mode", "") or "") if has_xml_evidence else ""
            ),
            "correo_enviado": correo_rx,
            "correo_clase": correo_clase,
            "recordatorios": "0",
            "observaciones": str(getattr(bo, "observaciones_recepcion", "") or ""),
            "observacion_descartes": "",
            "archivo_xml": str(getattr(bo, "descripcion", "") or ""),
            "archivo_xml_usado": str(getattr(bo, "descripcion", "") or ""),
            "observaciones_xml": obs_xml,
            "xml_clase": xml_clase,
            "numero_boleta_xml": fmt.format_folio(getattr(xml, "numero_boleta", None)) if xml else "",
            "fecha_boleta_xml": str(getattr(xml, "fecha_boleta", "") or "") if xml else "",
            "rut_emisor_xml": fmt.format_rut_cl(getattr(xml, "rut_emisor", "") if xml else "") or "",
            "rut_receptor_xml": fmt.format_rut_cl(getattr(xml, "rut_receptor", "") if xml else "") or "",
            "nombre_receptor_xml": "",
            "total_honorarios_xml": fmt.format_monto_cl(getattr(xml, "total_honorarios", None)) if xml else "",
            "liquido_honorarios_xml": fmt.format_monto_cl(getattr(xml, "liquido_honorarios", None)) if xml else "",
            "impuesto_honorarios_xml": fmt.format_monto_cl(getattr(xml, "impuesto_honorarios", None)) if xml else "",
            "descripcion_xml": str(getattr(xml, "descripcion_linea", "") or "") if xml else "",
            "correo_recepcion_enviado": correo_rx,
            "monto": fmt.format_monto_cl(getattr(bo, "monto_bruto", None)) or "",
        }
        out["rows"].append(row_payload)

    # KPIs de pagos desde snapshot JSONB del período
    try:
        import period_snapshots

        pagos_snap = period_snapshots.load_pagos_snapshot(year, month)
        if pagos_snap and pagos_snap.get("exists"):
            prow = pagos_snap.get("rows") or []
            out["pagos"]["sheet_exists"] = True
            out["pagos"]["total_rows"] = len(prow)
            for r in prow:
                correo = str(r.get("Correo Enviado") or "").strip().lower()
                if not correo:
                    out["pagos"]["pendiente"] += 1
                elif "error" in correo:
                    out["pagos"]["error"] += 1
                elif "omit" in correo:
                    out["pagos"]["omitido"] += 1
                elif "enviad" in correo:
                    out["pagos"]["enviado"] += 1
                else:
                    out["pagos"]["otro"] += 1
    except Exception:
        pass

    # Completar estado_pago en filas de avance (ya proyectado en boletas)
    try:
        for idx, (bo, *_rest) in enumerate(rows[:limit]):
            if idx < len(out["rows"]):
                ep = str(getattr(bo, "estado_pago", "") or "")
                if ep:
                    out["rows"][idx]["estado_pago"] = ep
                    out["rows"][idx]["observaciones_pago"] = str(
                        getattr(bo, "observaciones_pago", "") or ""
                    )
    except Exception:
        pass

    out["solicitud_exists"] = len(rows) > 0
    out["source"] = "postgresql"
    return out


def _check_item(item_id: str, label: str, ok: bool, message: str = "", *, blocking: bool = True) -> dict[str, Any]:
    return {
        "id": item_id,
        "label": label,
        "ok": ok,
        "blocking": blocking,
        "message": message,
    }


def prerequisite_checklist(stage_num: int, year: int | str, month: str) -> list[dict[str, Any]]:
    """Checklist detallado para la UI (bloqueo si algún ítem blocking falla)."""
    items: list[dict[str, Any]] = []
    month_path = _month_dir(year, month)
    month_exists = os.path.isdir(month_path)

    items.append(
        _check_item(
            "period_folder",
            "Carpeta del período existe",
            month_exists,
            "" if month_exists else f"No existe: {month_path}",
        )
    )

    if stage_num == 0:
        if month_exists:
            maestros = [
                f
                for f in os.listdir(month_path)
                if f.lower().endswith(".xlsx") and os.path.isfile(os.path.join(month_path, f))
            ]
            items.append(
                _check_item(
                    "maestro",
                    "Hay al menos un Excel maestro en la carpeta del mes",
                    len(maestros) > 0,
                    "Sube el archivo maestro (.xlsx) a la carpeta del mes.",
                )
            )
        root_xlsx = [
            f
            for f in os.listdir(config.RAIZ)
            if f.lower().endswith(".xlsx") and os.path.isfile(os.path.join(config.RAIZ, f))
        ]
        bd_ok = any("bd" in f.lower() or "docentes" in f.lower() for f in root_xlsx)
        items.append(
            _check_item(
                "bd_docentes",
                "BD-DOCENTES (o similar) en la raíz del proyecto",
                bd_ok,
                "Coloca BD-DOCENTES.xlsx en la raíz del repositorio.",
            )
        )
        return items

    if stage_num == 1:
        items.append(
            _check_item(
                "adjunto_ejemplo",
                "PDF de ejemplo para correos (EjemploEnvioBoleta.pdf)",
                os.path.isfile(config.ARCHIVO_ADJUNTO),
                f"No encontrado: {config.ARCHIVO_ADJUNTO}",
            )
        )

    solicitud = os.path.join(month_path, "Solicitud.xlsx")
    if stage_num in (1, 3, 4, 5, 6, 7, 8, 9, 10):
        items.append(
            _check_item(
                "solicitud",
                "Solicitud.xlsx en la carpeta del mes",
                os.path.isfile(solicitud),
                f"Ejecuta paso 0 o copia Solicitud.xlsx a {month_path}",
            )
        )

    if stage_num == 2:
        items.append(
            _check_item(
                "outlook_hint",
                "Outlook debe estar disponible en esta máquina",
                True,
                "",
                blocking=False,
            )
        )
        items.append(
            _check_item(
                "fechas_ui",
                "Indica fecha inicio y fin en el formulario",
                True,
                "",
                blocking=False,
            )
        )

    if stage_num == 8:
        map_path = os.path.join(month_path, "map_ip_cft.csv")
        map_ok = os.path.isfile(map_path)
        items.append(
            _check_item(
                "map_csv",
                "CSV de clasificación IP/CFT (map_ip_cft.csv)",
                map_ok,
                ""
                if map_ok
                else "Falta map_ip_cft.csv. Al ejecutar el paso 8 se genera desde Solicitud, o: python herramientas/generar_map_ip_cft.py --year … --month …",
                blocking=False,
            )
        )

    if stage_num in (5, 7):
        items.append(
            _check_item(
                "send_confirm",
                "Envío real requiere confirmación explícita en la UI",
                True,
                "",
                blocking=False,
            )
        )

    return items


def prerequisites_summary(checklist: list[dict[str, Any]]) -> dict[str, Any]:
    blocking = [c for c in checklist if c.get("blocking", True)]
    failed = [c for c in blocking if not c.get("ok")]
    return {
        "ok": len(failed) == 0,
        "message": failed[0]["message"] if failed else "",
        "failed_ids": [c["id"] for c in failed],
    }


def warnings_for_stage(stage_num: int, year: int | str, month: str) -> list[dict[str, str]]:
    warnings: list[dict[str, str]] = []
    month_path = _month_dir(year, month)
    if stage_num in stage_commands.EMAIL_STAGES:
        warnings.append(
            {
                "code": "EMAIL_STAGE",
                "message": "Etapa de correo: sin «Enviar correos reales» solo analiza/previsualiza.",
            }
        )
    if stage_num == 2:
        warnings.append(
            {
                "code": "OUTLOOK_COM",
                "message": "Requiere Outlook en ejecución en este equipo.",
            }
        )
    if stage_num == 8 and os.path.isdir(month_path):
        map_path = os.path.join(month_path, "map_ip_cft.csv")
        if not os.path.isfile(map_path):
            warnings.append(
                {
                    "code": "STEP8_MAP",
                    "message": "Paso 8: falta map_ip_cft.csv; al ejecutar se genera desde Solicitud (no uses Contabilidad_pagos.csv).",
                }
            )
    xml_count = 0
    if os.path.isdir(month_path):
        xml_count = len([f for f in os.listdir(month_path) if f.lower().endswith(".xml")])
    if stage_num == 3 and xml_count == 0:
        warnings.append(
            {
                "code": "NO_XML",
                "message": "No hay XML en la raíz del mes; ejecuta paso 2 primero.",
            }
        )
    return warnings


def estimated_outputs_for_stage(stage_num: int, year: int | str, month: str) -> list[dict[str, str]]:
    month_path = _month_dir(year, month)
    outputs: list[dict[str, str]] = [
        {"id": "job_log", "label": "Log del job (.log)"},
    ]
    if stage_num == 0:
        outputs.append({"id": "primary", "label": "Solicitud.xlsx generado"})
    elif stage_num == 10:
        outputs.append({"id": "revision_carpetas", "label": "revision_carpetas.xlsx"})
    elif stage_num == 9:
        outputs.append({"id": "resumen_agrupa", "label": "CSV resumen en logs_agrupa/"})
        outputs.append({"id": "primary", "label": "Carpetas IP/CFT por docente"})
    else:
        outputs.append({"id": "primary", "label": "Solicitud.xlsx actualizado"})
    if stage_num in stage_commands.EMAIL_STAGES:
        outputs.append({"id": "outbox", "label": "Entradas en outbox de correo (sqlite)"})
    return outputs


def period_summary(year: int | str, month: str) -> dict[str, Any]:
    """KPIs operativos del período (BD preferente, fallback Excel)."""
    month_path = _month_dir(year, month)
    summary: dict[str, Any] = {
        "year": int(year) if str(year).isdigit() else year,
        "month": month,
        "month_dir": month_path,
        "solicitud_exists": False,
        "total_rows": 0,
        "recibidos": 0,
        "no_recibidos": 0,
        "xml_files_in_month": 0,
        "pdf_files_in_month": 0,
    }
    if not os.path.isdir(month_path):
        # Aun sin carpeta local, intentamos devolver KPIs desde BD.
        if get_bool_setting("BH_READ_FROM_DB", True):
            return _period_summary_from_db(year, month) or summary
        return summary

    summary["xml_files_in_month"] = len(
        [f for f in os.listdir(month_path) if f.lower().endswith(".xml")]
    )
    summary["pdf_files_in_month"] = len(
        [f for f in os.listdir(month_path) if f.lower().endswith(".pdf")]
    )
    if get_bool_setting("BH_READ_FROM_DB", True):
        from_db = _period_summary_from_db(year, month)
        if from_db is not None:
            # Conserva métricas de archivos en carpeta para contexto operativo.
            from_db["xml_files_in_month"] = summary["xml_files_in_month"]
            from_db["pdf_files_in_month"] = summary["pdf_files_in_month"]
            return from_db

    solicitud = os.path.join(month_path, "Solicitud.xlsx")
    if not os.path.isfile(solicitud):
        return summary

    summary["solicitud_exists"] = True
    try:
        import pandas as pd

        df = pd.read_excel(solicitud, sheet_name=0, engine="openpyxl")
        summary["total_rows"] = int(len(df))
        if "Estado_Recepcion" in df.columns:
            vc = df["Estado_Recepcion"].astype(str).str.upper().str.strip()
            summary["recibidos"] = int(vc.isin({"RECIBIDO", "RECIBIDO CON ERROR"}).sum())
            summary["no_recibidos"] = int((vc == "NO RECIBIDO").sum())
    except Exception as e:
        summary["read_error"] = str(e)
    return summary


def _period_summary_from_db(year: int | str, month: str) -> dict[str, Any] | None:
    """KPIs de período directamente desde PostgreSQL."""
    from sqlalchemy import func, select
    from sqlalchemy.exc import SQLAlchemyError

    from db.models import Boleta, BoletaXmlData, Periodo
    from db.session import SessionLocal

    month_norm = str(month or "").strip().capitalize()
    try:
        year_int = int(year)
    except (TypeError, ValueError):
        return None
    month_path = _month_dir(year, month)
    base = {
        "year": year_int,
        "month": month,
        "month_dir": month_path,
        "solicitud_exists": os.path.isfile(os.path.join(month_path, "Solicitud.xlsx")),
        "total_rows": 0,
        "recibidos": 0,
        "no_recibidos": 0,
        "xml_files_in_month": 0,
        "pdf_files_in_month": 0,
    }
    try:
        with SessionLocal() as session:
            periodo = session.execute(
                select(Periodo).where(Periodo.anio == year_int, Periodo.mes_nombre == month_norm)
            ).scalar_one_or_none()
            if periodo is None:
                return None
            total = session.execute(
                select(func.count(Boleta.id)).where(Boleta.periodo_id == periodo.id)
            ).scalar_one()
            recibidos = session.execute(
                select(func.count(Boleta.id)).where(
                    Boleta.periodo_id == periodo.id,
                    func.coalesce(Boleta.estado_recepcion, "").in_(["RECIBIDO", "RECIBIDO CON ERROR"]),
                )
            ).scalar_one()
            no_recibidos = session.execute(
                select(func.count(Boleta.id)).where(
                    Boleta.periodo_id == periodo.id,
                    func.coalesce(Boleta.estado_recepcion, "") == "NO RECIBIDO",
                )
            ).scalar_one()
            xml_count = session.execute(
                select(func.count(BoletaXmlData.id))
                .join(Boleta, BoletaXmlData.boleta_id == Boleta.id)
                .where(Boleta.periodo_id == periodo.id)
            ).scalar_one()
            total_rows = int(total or 0)
            return {
                **base,
                "total_rows": total_rows,
                "recibidos": int(recibidos or 0),
                "no_recibidos": int(no_recibidos or 0),
                "xml_files_in_month": int(xml_count or 0),
                "solicitud_exists": total_rows > 0,
            }
    except SQLAlchemyError as exc:
        base["read_error"] = str(exc)
        return base


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        import pandas as pd

        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def _solicitud_field(solicitud_row: dict[str, Any] | None, *names: str) -> str:
    """Lee un dato de la fila Solicitud guardada en BD (NAME, email, sede, etc.)."""
    if not isinstance(solicitud_row, dict):
        return ""
    for name in names:
        text = _cell_str(solicitud_row.get(name))
        if text:
            return text
    return ""


def _classify_mail_flag(value: Any) -> str:
    text = _cell_str(value)
    if not text:
        return "pendiente"
    low = text.lower()
    if "❌" in text or "error" in low or "inválido" in low or "invalido" in low:
        return "error"
    if "omitido" in low:
        return "omitido"
    if "✅" in text or "enviado" in low:
        return "enviado"
    return "otro"


def _classify_xml_obs(value: Any) -> str:
    text = _cell_str(value)
    if not text:
        return "pendiente"
    low = text.upper()
    if "DATOS EXTRAIDOS OK" in low or low.endswith(" OK") or "EXTRAIDOS OK" in low:
        return "ok"
    return "observacion"


def _read_solicitud_workbook(path: str):
    """Lee hojas útiles de Solicitud.xlsx; si Excel lo tiene abierto, copia a temp."""
    import shutil
    import tempfile

    import pandas as pd
    from openpyxl import load_workbook
    from schema_validator import find_sheet

    def _load(src: str):
        wb = load_workbook(src, read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
        finally:
            wb.close()
        solicitud_sheet = find_sheet(sheet_names, "Solicitud") or sheet_names[0]
        df = pd.read_excel(src, sheet_name=solicitud_sheet, engine="openpyxl")
        pagos_sheet = find_sheet(sheet_names, "Pagos")
        df_pagos = None
        if pagos_sheet:
            df_pagos = pd.read_excel(src, sheet_name=pagos_sheet, engine="openpyxl")
        return df, sheet_names, solicitud_sheet, pagos_sheet, df_pagos

    try:
        return _load(path)
    except PermissionError:
        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            shutil.copy2(path, tmp)
            return _load(tmp)
        finally:
            try:
                os.remove(tmp)
            except OSError:
                pass


def excel_avance(year: int | str, month: str, *, row_limit: int = 500) -> dict[str, Any]:
    """Avance vivo leído de Solicitud.xlsx (sin abrir el archivo en Excel)."""
    if get_bool_setting("BH_READ_FROM_DB", True):
        from_db = _excel_avance_from_db(year, month, row_limit=row_limit)
        if from_db is not None:
            return from_db
        # Modo fuente de verdad BD: evita fallback silencioso a Excel.
        month_path = _month_dir(year, month)
        return {
            "year": int(year) if str(year).isdigit() else year,
            "month": month,
            "month_dir": month_path,
            "solicitud_path": os.path.join(month_path, "Solicitud.xlsx"),
            "solicitud_exists": os.path.isfile(os.path.join(month_path, "Solicitud.xlsx")),
            "sheets": [],
            "solicitud_sheet": None,
            "total_rows": 0,
            "recepcion": {
                "recibido": 0,
                "recibido_con_error": 0,
                "no_recibido": 0,
                "pendiente": 0,
                "otro": 0,
            },
            "correo_solicitud": {"enviado": 0, "omitido": 0, "error": 0, "pendiente": 0, "otro": 0},
            "recordatorios": {"con_recordatorio": 0, "total_envios": 0},
            "xml_extract": {"ok": 0, "observacion": 0, "pendiente": 0, "con_archivo": 0},
            "archivos_mes": {"xml": 0, "pdf": 0},
            "pagos": {
                "sheet_exists": False,
                "total_rows": 0,
                "enviado": 0,
                "pendiente": 0,
                "error": 0,
                "omitido": 0,
                "otro": 0,
            },
            "rows": [],
            "rows_truncated": False,
            "read_error": "BH_READ_FROM_DB=1 y no se pudo leer estado canónico desde PostgreSQL.",
            "mtime": None,
        }
    month_path = _month_dir(year, month)
    out: dict[str, Any] = {
        "year": int(year) if str(year).isdigit() else year,
        "month": month,
        "month_dir": month_path,
        "solicitud_path": os.path.join(month_path, "Solicitud.xlsx"),
        "solicitud_exists": False,
        "sheets": [],
        "solicitud_sheet": None,
        "total_rows": 0,
        "recepcion": {
            "recibido": 0,
            "recibido_con_error": 0,
            "no_recibido": 0,
            "pendiente": 0,
            "otro": 0,
        },
        "correo_solicitud": {"enviado": 0, "omitido": 0, "error": 0, "pendiente": 0, "otro": 0},
        "recordatorios": {"con_recordatorio": 0, "total_envios": 0},
        "xml_extract": {"ok": 0, "observacion": 0, "pendiente": 0, "con_archivo": 0},
        "archivos_mes": {"xml": 0, "pdf": 0},
        "pagos": {
            "sheet_exists": False,
            "total_rows": 0,
            "enviado": 0,
            "pendiente": 0,
            "error": 0,
            "omitido": 0,
            "otro": 0,
        },
        "rows": [],
        "rows_truncated": False,
        "read_error": None,
        "mtime": None,
    }

    if not os.path.isdir(month_path):
        return out

    out["archivos_mes"]["xml"] = len(
        [f for f in os.listdir(month_path) if f.lower().endswith(".xml")]
    )
    out["archivos_mes"]["pdf"] = len(
        [f for f in os.listdir(month_path) if f.lower().endswith(".pdf")]
    )

    solicitud = out["solicitud_path"]
    if not os.path.isfile(solicitud):
        return out

    out["solicitud_exists"] = True
    try:
        out["mtime"] = datetime.fromtimestamp(os.path.getmtime(solicitud)).isoformat(timespec="seconds")
    except OSError:
        pass

    try:
        df, sheet_names, sheet, pagos_sheet, df_pagos = _read_solicitud_workbook(solicitud)
        out["sheets"] = sheet_names
        out["solicitud_sheet"] = sheet
        out["total_rows"] = int(len(df))

        if "Estado_Recepcion" in df.columns:
            for _, row in df.iterrows():
                _estado_rx, estado_ef, _glosa_ok, _glosa_xml, _glosa_mode = _recepcion_efectivo_fila(
                    row, df.columns, month_path
                )
                _inc_recepcion_kpi(out["recepcion"], estado_ef)

        if "Correo Enviado" in df.columns:
            for raw in df["Correo Enviado"]:
                out["correo_solicitud"][_classify_mail_flag(raw)] += 1

        if "Recordatorios Enviados" in df.columns:
            for raw in df["Recordatorios Enviados"]:
                text = _cell_str(raw)
                if not text:
                    continue
                try:
                    n = int(float(text.replace(",", ".")))
                except ValueError:
                    digits = "".join(ch for ch in text if ch.isdigit())
                    n = int(digits) if digits else 0
                if n > 0:
                    out["recordatorios"]["con_recordatorio"] += 1
                    out["recordatorios"]["total_envios"] += n

        if "Observaciones_XML" in df.columns:
            for raw in df["Observaciones_XML"]:
                out["xml_extract"][_classify_xml_obs(raw)] += 1
        else:
            out["xml_extract"]["pendiente"] = out["total_rows"]

        if "archivo_xml" in df.columns:
            out["xml_extract"]["con_archivo"] = int(
                sum(1 for raw in df["archivo_xml"] if _cell_str(raw))
            )

        if pagos_sheet and df_pagos is not None:
            out["pagos"]["sheet_exists"] = True
            out["pagos"]["total_rows"] = int(len(df_pagos))
            if "Correo Enviado" in df_pagos.columns:
                for raw in df_pagos["Correo Enviado"]:
                    out["pagos"][_classify_mail_flag(raw)] += 1
            else:
                out["pagos"]["pendiente"] = out["pagos"]["total_rows"]

        if row_limit > 0:
            limit = min(int(row_limit), len(df))
            out["rows_truncated"] = len(df) > limit
            rows: list[dict[str, Any]] = []

            for idx in range(limit):
                row = df.iloc[idx]
                correo_raw = row.get("Correo Enviado") if "Correo Enviado" in df.columns else ""
                obs_xml_raw = (
                    row.get("Observaciones_XML") if "Observaciones_XML" in df.columns else ""
                )
                estado_rx, estado_rx_efectivo, glosa_xml_coincide, glosa_xml, glosa_match_mode = _recepcion_efectivo_fila(
                    row, df.columns, month_path
                )
                glosa_pedida = _excel_row_str(row, df.columns, "GLOSA")
                rows.append(
                    {
                        "row": int(idx) + 2,
                        "emplid": fmt.format_rut_cl(row.get("EMPLID")) or _excel_row_str(row, df.columns, "EMPLID"),
                        "rut_sin_dv": fmt.format_rut_sin_dv(row.get("RUT_SIN_DV")) or _excel_row_str(row, df.columns, "RUT_SIN_DV"),
                        "name": _excel_row_str(row, df.columns, "NAME"),
                        "sede": _excel_row_str(row, df.columns, "SEDE"),
                        "location": _excel_row_str(row, df.columns, "LOCATION"),
                        "email": _excel_row_str(row, df.columns, "Email_Docente"),
                        "email_dp": _excel_row_str(row, df.columns, "Email_DP"),
                        "rut_razon": fmt.format_rut_cl(row.get("RUT RAZON")) or _excel_row_str(row, df.columns, "RUT RAZON"),
                        "nombre_razon": _excel_row_str(row, df.columns, "NOMBRE RAZON"),
                        "direccion_razon": _excel_row_str(row, df.columns, "DireccionRazon"),
                        "glosa": glosa_pedida,
                        "provisionado": _es_glosa_provisionado(glosa_pedida),
                        "estado_recepcion": estado_rx,
                        "estado_recepcion_efectivo": estado_rx_efectivo,
                        "glosa_xml_coincide": glosa_xml_coincide,
                        "glosa_match_mode": glosa_match_mode,
                        "correo_enviado": _excel_row_str(row, df.columns, "Correo Enviado"),
                        "correo_clase": _classify_mail_flag(correo_raw),
                        "recordatorios": _excel_row_str(row, df.columns, "Recordatorios Enviados"),
                        "observaciones": _excel_row_str(row, df.columns, "Observaciones"),
                        "observacion_descartes": _excel_row_str(row, df.columns, "Observacion_Descartes"),
                        "archivo_xml": _excel_row_str(row, df.columns, "archivo_xml"),
                        "archivo_xml_usado": _excel_row_str(row, df.columns, "Archivo_XML_Usado"),
                        "observaciones_xml": _excel_row_str(row, df.columns, "Observaciones_XML"),
                        "xml_clase": _classify_xml_obs(obs_xml_raw),
                        "numero_boleta_xml": fmt.format_folio(row.get("numeroBoleta_XML")) or _excel_row_str(row, df.columns, "numeroBoleta_XML"),
                        "fecha_boleta_xml": _excel_row_str(row, df.columns, "fechaBoleta_XML"),
                        "rut_emisor_xml": fmt.format_rut_cl(row.get("rutEmisorCompleto_XML")) or _excel_row_str(row, df.columns, "rutEmisorCompleto_XML"),
                        "rut_receptor_xml": fmt.format_rut_cl(row.get("rutReceptorCompleto_XML")) or _excel_row_str(row, df.columns, "rutReceptorCompleto_XML"),
                        "nombre_receptor_xml": _excel_row_str(row, df.columns, "nombreReceptor_XML"),
                        "total_honorarios_xml": fmt.format_monto_cl(row.get("totalHonorarios_XML")) or _excel_row_str(row, df.columns, "totalHonorarios_XML"),
                        "liquido_honorarios_xml": fmt.format_monto_cl(row.get("liquidoHonorarios_XML")) or _excel_row_str(row, df.columns, "liquidoHonorarios_XML"),
                        "impuesto_honorarios_xml": fmt.format_monto_cl(row.get("impuestoHonorarios_XML")) or _excel_row_str(row, df.columns, "impuestoHonorarios_XML"),
                        "descripcion_xml": glosa_xml,
                        "correo_recepcion_enviado": _excel_row_str(row, df.columns, "Correo_Recepcion_Enviado"),
                        "monto": fmt.format_monto_cl(row.get("CUS_TOT_HON")) or _excel_row_str(row, df.columns, "CUS_TOT_HON"),
                    }
                )
            out["rows"] = rows
    except Exception as e:
        out["read_error"] = str(e)

    return out


def _last_job_for_stage(jobs: list[dict[str, Any]], stage_num: int) -> dict[str, Any] | None:
    filtered = [j for j in jobs if j.get("stage_num") == stage_num]
    if not filtered:
        return None
    filtered.sort(key=lambda j: j.get("created_at", ""), reverse=True)
    return filtered[0]


def _status_from_last_execution(last: dict[str, Any] | None) -> str | None:
    if not last:
        return None
    st = str(last.get("status", "")).lower()
    if st == "running":
        return "RUNNING"
    if st == "failed":
        return "ERROR"
    if st == "success":
        return "OK"
    # Jobs API antiguos sin status; un log de disco "unknown" no cuenta como paso hecho.
    if st == "unknown" and str(last.get("source") or "") == "api":
        return "OK"
    return None


def _solicitud_has_resumen_boletas(year: int | str, month: str) -> bool:
    """True si Solicitud.xlsx ya tiene la hoja del informe (paso 6)."""
    path = os.path.join(_month_dir(year, month), "Solicitud.xlsx")
    if not os.path.isfile(path):
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=False)
        names = {str(n).strip().lower() for n in wb.sheetnames}
        wb.close()
        return "resumen boletas" in names
    except Exception:
        return False


def ui_status_for_stage(
    stage_num: int,
    year: int | str,
    month: str,
    *,
    jobs: list[dict[str, Any]] | None = None,
    running_job: dict[str, Any] | None = None,
    last_execution: dict[str, Any] | None = None,
) -> str:
    if running_job and running_job.get("stage_num") == stage_num:
        return "RUNNING"
    last = last_execution
    if not last:
        job = _last_job_for_stage(jobs or [], stage_num)
        if job:
            last = ops_execution_history._api_job_summary(job)
    from_last = _status_from_last_execution(last)
    if from_last:
        return from_last
    if stage_num == 0:
        solicitud = os.path.join(_month_dir(year, month), "Solicitud.xlsx")
        if os.path.isfile(solicitud):
            return "OK"
    # Paso 6 no escribe log histórico por defecto; la hoja Resumen Boletas es la evidencia.
    if stage_num == 6 and _solicitud_has_resumen_boletas(year, month):
        return "OK"
    checklist = prerequisite_checklist(stage_num, year, month)
    if not prerequisites_summary(checklist)["ok"]:
        return "BLOCKED"
    return "READY"


def inbox_gaps_scan(
    year: int | str,
    month: str,
    *,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
) -> dict[str, Any]:
    """Huecos Inbox↔carpeta para filas NO RECIBIDO (detector Maass)."""
    return inbox_gaps.detectar_huecos_inbox(
        year,
        month,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )


def period_overview(
    year: int,
    month: str,
    *,
    jobs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    jobs = jobs or []
    period_jobs = [
        j
        for j in jobs
        if str(j.get("year")) == str(year) and str(j.get("month")) == month
    ]
    running = next((j for j in period_jobs if j.get("status") == "running"), None)
    history_by_stage = ops_execution_history.last_executions_by_stage(
        year, month, api_jobs=jobs
    )

    stages_out: list[dict[str, Any]] = []
    for meta in stage_commands.list_stages_metadata():
        sn = meta["stage_num"]
        checklist = prerequisite_checklist(sn, year, month)
        prereq = prerequisites_summary(checklist)
        last = history_by_stage.get(sn)
        stages_out.append(
            {
                **meta,
                "ui_status": ui_status_for_stage(
                    sn,
                    year,
                    month,
                    jobs=period_jobs,
                    running_job=running,
                    last_execution=last,
                ),
                "prerequisites": prereq,
                "checklist": checklist,
                "warnings": warnings_for_stage(sn, year, month),
                "estimated_outputs": estimated_outputs_for_stage(sn, year, month),
                "last_job": last,
            }
        )

    try:
        import email_outbox

        outbox_stats = email_outbox.stats_by_status()
    except Exception:
        outbox_stats = {}

    kpis = period_summary(year, month)
    period_status = None
    try:
        import period_policy

        period_status = period_policy.get_period_status(year, month)
    except Exception:
        period_status = None

    try:
        import sync_status

        sync_status_out = sync_status.period_sync_status(year, month)
    except Exception as exc:
        sync_status_out = {"status": "unknown", "message": f"No se pudo evaluar sync_status: {exc}", "details": {}}

    return {
        "period": {"year": year, "month": month, "status": period_status or "abierto"},
        "kpis": kpis,
        "stages": stages_out,
        "running_job": (
            {"id": running["id"], "stage_num": running.get("stage_num"), "type": running.get("type")}
            if running
            else None
        ),
        "outbox_stats": outbox_stats,
        "sync_status": sync_status_out,
        "recommendation": recommend_next_action(
            stages_out,
            kpis=kpis,
            running_job=running,
            outbox_stats=outbox_stats,
            period_status=period_status,
        ),
    }


def recommend_next_action(
    stages: list[dict[str, Any]],
    *,
    kpis: dict[str, Any],
    running_job: dict[str, Any] | None = None,
    outbox_stats: dict[str, int] | None = None,
    period_status: str | None = None,
) -> dict[str, Any]:
    """Sugerencia del siguiente paso operativo (misma lógica que la UI)."""
    outbox_stats = outbox_stats or {}
    try:
        import period_policy

        if period_policy.is_closed_status(period_status):
            return {
                "kind": "review",
                "stage_num": None,
                "title": "Período cerrado",
                "message": (
                    "Este período está cerrado en la base de datos. "
                    "La API no permite jobs ni sesiones interactivas. "
                    "Selecciona un mes abierto o usa la consola con supervisión manual."
                ),
                "action_label": "Cambiar período",
            }
    except Exception:
        pass

    sorted_stages = sorted(stages, key=lambda s: int(s["stage_num"]))

    if running_job:
        sn = int(running_job.get("stage_num", -1))
        return {
            "kind": "wait",
            "stage_num": sn,
            "title": "Job en ejecución",
            "message": f"Espera a que termine el paso {sn} (job {running_job.get('id', '')}).",
            "action_label": "Ver seguimiento",
        }

    pending_outbox = int(outbox_stats.get("pending", 0))
    if pending_outbox > 0:
        return {
            "kind": "outbox",
            "stage_num": None,
            "title": "Correos pendientes en outbox",
            "message": f"Hay {pending_outbox} envío(s) pending. Revisa pestaña Avanzado → Outbox o ejecuta dispatch COM.",
            "action_label": "Ir a Avanzado",
        }

    failed = next(
        (s for s in sorted_stages if s.get("ui_status") == "ERROR" and s.get("enabled_for_api")),
        None,
    )
    if failed:
        sn = int(failed["stage_num"])
        return _recommend_run(
            failed,
            title=f"Reintentar paso {sn}",
            message=f"La última ejecución del paso {sn} falló. Revisa logs y vuelve a ejecutar.",
        )

    if int(kpis.get("total_rows") or 0) == 0:
        step0 = next((s for s in sorted_stages if s.get("stage_num") == 0), None)
        if step0 and step0.get("ui_status") != "OK":
            return _recommend_run(
                step0,
                title="Generar Solicitud",
                message="Aún no hay solicitudes en este período. Empieza por el paso 0.",
            )

    blocked = next((s for s in sorted_stages if s.get("ui_status") == "BLOCKED"), None)
    if blocked:
        sn = int(blocked["stage_num"])
        prereq = blocked.get("prerequisites") or {}
        msg = prereq.get("message") or "Completa los requisitos del checklist."
        return {
            "kind": "fix",
            "stage_num": sn,
            "title": f"Desbloquear paso {sn}",
            "message": msg,
            "action_label": f"Ver paso {sn}",
        }

    def _is_ready_stage(s: dict[str, Any]) -> bool:
        return s.get("ui_status") == "READY" and s.get("enabled_for_api")

    ready = next((s for s in sorted_stages if _is_ready_stage(s)), None)
    if ready and int(ready.get("stage_num", -1)) == 0 and int(kpis.get("total_rows") or 0) > 0:
        ready = next(
            (s for s in sorted_stages if _is_ready_stage(s) and int(s["stage_num"]) > 0),
            None,
        )

    no_recibidos = int(kpis.get("no_recibidos") or 0)
    stage3 = next((s for s in sorted_stages if int(s.get("stage_num", -1)) == 3), None)
    stage3_ok = bool(stage3 and stage3.get("ui_status") == "OK")

    def _reminders_recommendation() -> dict[str, Any]:
        return {
            "kind": "reminders",
            "stage_num": 1,
            "title": "Recordatorios a pendientes",
            "message": (
                f"Hay {no_recibidos} fila(s) NO RECIBIDO. "
                "Vuelve al paso 1 en modo solo recordatorios (no reenvía solicitudes originales)."
            ),
            "action_label": "Paso 1 · solo recordatorios",
            "params": {"reminders_only": True},
        }

    # Tras validar (paso 3 OK), si ya no hay trabajo de recepción 2–5 pendiente
    # y aún faltan boletas → priorizar recordatorios antes de pasos ≥6 o “completo”.
    if no_recibidos > 0 and stage3_ok:
        inbound_ready = (
            ready is not None and int(ready.get("stage_num", -1)) in (2, 3, 4, 5)
        )
        if not inbound_ready:
            return _reminders_recommendation()

    if ready:
        sn = int(ready["stage_num"])
        stage1 = next((s for s in sorted_stages if int(s.get("stage_num", -1)) == 1), None)
        stage1_ok = bool(stage1 and stage1.get("ui_status") == "OK")
        xml_count = int(kpis.get("xml_files_in_month") or 0)
        if sn == 2 and stage1_ok and xml_count == 0:
            return _recommend_run(
                ready,
                title="Cuando lleguen boletas: paso 2",
                message=(
                    "Las solicitudes ya se enviaron. Cuando empiecen a llegar boletas por correo, "
                    "ejecuta el paso 2. Por defecto usa todo el mes; no acotes fechas de más "
                    "(si llega algo fuera del rango, no se baja)."
                ),
            )
        hint = _stage_run_hint(ready, kpis)
        return _recommend_run(
            ready,
            title=f"Siguiente: paso {sn}",
            message=hint,
        )

    api_stages = [s for s in sorted_stages if s.get("enabled_for_api")]

    def _effectively_complete(s: dict[str, Any]) -> bool:
        if int(s["stage_num"]) == 0 and int(kpis.get("total_rows") or 0) > 0:
            return s.get("ui_status") in ("OK", "READY")
        return s.get("ui_status") == "OK"

    if api_stages and all(_effectively_complete(s) for s in api_stages):
        if no_recibidos > 0 and stage3_ok:
            return _reminders_recommendation()
        return {
            "kind": "complete",
            "stage_num": 10,
            "title": "Pipeline API al día",
            "message": "Todos los pasos 0–10 muestran última ejecución OK. Revisa carpeta/revisión o consola si falta algo manual.",
            "action_label": "Ver paso 10",
        }

    if no_recibidos > 0 and stage3_ok:
        return _reminders_recommendation()

    return {
        "kind": "review",
        "stage_num": 0,
        "title": "Revisar estado",
        "message": "Revisa el listado de pasos en la barra lateral.",
        "action_label": "Ver paso 0",
    }


def _recommend_run(stage: dict[str, Any], *, title: str, message: str) -> dict[str, Any]:
    sn = int(stage["stage_num"])
    extra = ""
    if stage.get("is_email_stage"):
        extra = " En pasos de correo, marca «Enviar correos reales» solo si corresponde."
    return {
        "kind": "run",
        "stage_num": sn,
        "title": title,
        "message": message + extra,
        "action_label": f"Ir a paso {sn}",
    }


def _stage_run_hint(stage: dict[str, Any], kpis: dict[str, Any]) -> str:
    sn = int(stage["stage_num"])
    desc = str(stage.get("description", ""))
    if sn == 2:
        xml = int(kpis.get("xml_files_in_month") or 0)
        if xml == 0:
            return (
                "Bajar boletas desde Outlook. Deja el rango del mes completo salvo que sepas "
                "que debes ampliarlo."
            )
        return f"Extraer boletas del correo ({desc}). Hay {xml} XML en carpeta."
    if sn == 3:
        rec = int(kpis.get("recibidos", 0))
        xml = int(kpis.get("xml_files_in_month", 0))
        return f"Validar recepción ({desc}). Hay {xml} XML y {rec} filas recibidas en Excel."
    if sn == 4:
        return f"Volcar datos XML al Excel ({desc})."
    if sn in (5, 7):
        return f"{desc}. Sin envío real solo previsualiza/analiza."
    if sn == 8:
        return f"{desc}. Usa map_ip_cft.csv en la carpeta del mes (generar con herramienta)."
    return desc or f"Ejecutar paso {sn}."


def _safe_under_roots(path: str) -> bool:
    path = os.path.abspath(path)
    roots = [os.path.abspath(config.RAIZ), os.path.abspath(os.path.join(config.RAIZ, ".state"))]
    return any(path == r or path.startswith(r + os.sep) for r in roots)


def discover_period_artifacts(stage_num: int, year: int | str, month: str) -> list[dict[str, Any]]:
    """Artefactos conocidos en disco para un período/paso."""
    month_path = _month_dir(year, month)
    found: list[dict[str, Any]] = []

    def add(aid: str, label: str, path: str, kind: str) -> None:
        if path and os.path.isfile(path) and _safe_under_roots(path):
            found.append(
                {
                    "id": aid,
                    "label": label,
                    "path": path,
                    "filename": os.path.basename(path),
                    "kind": kind,
                    "exists": True,
                    "size_bytes": os.path.getsize(path),
                }
            )

    solicitud = os.path.join(month_path, "Solicitud.xlsx")
    add("solicitud", "Solicitud.xlsx", solicitud, "xlsx")
    revision = os.path.join(month_path, "revision_carpetas.xlsx")
    add("revision_carpetas", "revision_carpetas.xlsx", revision, "xlsx")

    agrupa_dir = os.path.join(month_path, "logs_agrupa")
    if os.path.isdir(agrupa_dir):
        csvs = sorted(glob.glob(os.path.join(agrupa_dir, "resumen_agrupa_*.csv")))
        if csvs:
            add("resumen_agrupa", "Resumen agrupación (último CSV)", csvs[-1], "csv")

    return found


def artifacts_for_job(job: dict[str, Any]) -> list[dict[str, Any]]:
    """Artefactos descargables asociados a un job."""
    artifacts: list[dict[str, Any]] = []
    job_id = job.get("id", "")

    def add(aid: str, label: str, path: str, kind: str) -> None:
        if not path or not os.path.isfile(path):
            return
        if not _safe_under_roots(path):
            return
        artifacts.append(
            {
                "id": aid,
                "label": label,
                "path": path,
                "filename": os.path.basename(path),
                "kind": kind,
                "exists": True,
                "size_bytes": os.path.getsize(path),
                "download_url": f"/operations/jobs/{job_id}/artifacts/{aid}",
            }
        )

    log_path = job.get("log_path")
    add("job_log", f"Log del job ({job_id})", log_path, "log")

    stage_num = int(job.get("stage_num", 0))
    year = job.get("year")
    month = job.get("month")
    params = job.get("params") or {}

    primary = job.get("output_path")
    if not primary and year and month:
        primary = stage_commands.primary_output_for_stage(stage_num, year, month, params)
    if primary:
        add("primary", os.path.basename(primary), primary, "xlsx")

    if year and month:
        for art in discover_period_artifacts(stage_num, year, month):
            if art["id"] == "primary" and any(a["id"] == "primary" for a in artifacts):
                continue
            if art["id"] not in {a["id"] for a in artifacts}:
                art = dict(art)
                art["download_url"] = f"/operations/jobs/{job_id}/artifacts/{art['id']}"
                artifacts.append(art)

    return artifacts


def resolve_job_artifact_path(job: dict[str, Any], artifact_id: str) -> str | None:
    for art in artifacts_for_job(job):
        if art["id"] == artifact_id and art.get("exists"):
            return art["path"]
    return None
