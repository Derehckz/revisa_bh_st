"""Etapa 3 — validación recepción PDF/XML vs Solicitud."""
from __future__ import annotations

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import config
import schema_validator
import utils
from db import boleta_repository, docente_repository, file_repository
from db.key_builder import build_boleta_key
from interaction.exceptions import SessionCancelled
from interaction.port import InteractionPort
from stages.context import Stage3Context
from stages.stage3 import revision_core as core

RAIZ = core.RAIZ


class Stage3Service:
    def run(self, ctx: Stage3Context, ui: InteractionPort) -> dict:
        ui.header("Validación de recepción PDF/XML", "Solicitud vs archivos bhe_ en carpeta del mes")

        try:
            año, mes = utils.resolve_año_mes(RAIZ, ctx.year, ctx.month)
        except ValueError as e:
            ui.log(str(e), level="error")
            return {"ok": False}

        ruta_mes = os.path.join(RAIZ, año, mes)
        mes_num = config.MESES_ES.index(mes) + 1 if mes in config.MESES_ES else 0

        ruta_logs = os.path.join(ruta_mes, "logs_revision")
        os.makedirs(ruta_logs, exist_ok=True)
        ruta_log_file = os.path.join(
            ruta_logs, datetime.now().strftime("revision_%Y%m%d_%H%M%S.log")
        )
        utils.configurar_logging(ruta_log_file)

        ruta_excel = os.path.join(ruta_mes, "Solicitud.xlsx")
        if not os.path.isfile(ruta_excel):
            ui.log(f"No se encontró Solicitud.xlsx en {ruta_mes}", level="error")
            return {"ok": False}

        ruta_bd = os.path.join(config.RAIZ, "BD-DOCENTES.xlsx")
        if os.path.isfile(ruta_bd):
            stats_doc = docente_repository.sync_docentes_from_excel(ruta_bd)
            ui.log(
                f"BD docentes: +{stats_doc['inserted']} nuevos, {stats_doc['updated']} actualizados",
                level="info",
            )
        else:
            ui.log("BD-DOCENTES.xlsx no encontrada; catálogo sin cambios.", level="warning")

        xml_count = len(
            [
                f
                for f in os.listdir(ruta_mes)
                if f.lower().endswith(".xml") and f.lower().startswith(config.PREFIJO.lower())
            ]
        ) if os.path.isdir(ruta_mes) else 0

        ui.table(
            "Contexto",
            [
                ("Raíz", RAIZ),
                ("Período", f"{mes} {año}"),
                ("Carpeta mes", ruta_mes),
                ("Excel", ruta_excel),
                ("XML en carpeta (bhe_)", str(xml_count)),
                ("Logs", ruta_log_file),
            ],
        )
        ui.emit("scan.ready", {"xml_files": xml_count, "month_dir": ruta_mes})

        if not ui.confirm_yes_no(
            "Continuar",
            "¿Validar recepción comparando planilla con PDF/XML del mes?",
            default=False,
        ):
            ui.log("Cancelado por el usuario.", level="warning")
            return {"ok": False, "cancelled": True}

        wb = load_workbook(ruta_excel, read_only=True)
        hojas = wb.sheetnames
        wb.close()

        if ctx.sheet and ctx.sheet in hojas:
            hoja = ctx.sheet
        elif len(hojas) == 1:
            hoja = hojas[0]
        elif utils.is_non_interactive():
            hoja = utils.pick_excel_sheet(hojas)
        else:
            hoja = str(
                ui.choose_option(
                    "Hoja Excel",
                    "Seleccione la hoja a validar",
                    hojas,
                    icon="📄",
                )
            )

        df = pd.read_excel(ruta_excel, sheet_name=hoja, engine="openpyxl")

        canonical_errors, canonical_warnings = schema_validator.validate_for_stage(
            df, "stage3_validacion_recepcion"
        )
        for w in canonical_warnings:
            logging.warning(f"[stage3] WARN {w}")
            ui.log(f"[schema] {w}", level="warning")
        for e in canonical_errors:
            logging.error(f"[stage3] ERROR {e}")
            ui.log(f"[schema] {e}", level="error")
        if canonical_errors and ctx.strict:
            ui.log("Validación estricta: abortando.", level="error")
            return {"ok": False}

        if ctx.supervised and not ui.confirm_yes_no(
            "Procesar filas",
            f"¿Procesar {len(df)} filas de la hoja «{hoja}»?",
            default=False,
        ):
            return {"ok": False, "cancelled": True}

        try:
            df_actualizado, stats = core.procesar_filas(df, ruta_mes, ui)
        except SessionCancelled:
            ui.log("Sesión cancelada.", level="warning")
            return {"ok": False, "cancelled": True}

        periodo_id = None
        if mes_num > 0:
            periodo_id = file_repository.get_or_create_periodo(
                anio=int(año),
                mes_num=mes_num,
                mes_nombre=mes,
            )
        for _, fila in df_actualizado.iterrows():
            boleta_repository.upsert_boleta_recepcion(
                periodo_id=periodo_id,
                boleta_key=build_boleta_key(fila.to_dict()),
                emplid=str(fila.get("EMPLID", "")).strip()
                if "EMPLID" in df_actualizado.columns
                else None,
                rut_sin_dv=str(fila.get("RUT_SIN_DV", "")).strip()
                if "RUT_SIN_DV" in df_actualizado.columns
                else None,
                rut_razon=str(fila.get("RUT RAZON", "")).strip()
                if "RUT RAZON" in df_actualizado.columns
                else None,
                estado_recepcion=str(fila.get("Estado_Recepcion", "")).strip() or None,
                observaciones_recepcion=str(fila.get("Observaciones", "")).strip() or None,
                glosa=str(fila.get("GLOSA", "")).strip()
                if "GLOSA" in df_actualizado.columns
                else None,
                monto_bruto=fila.get("CUS_TOT_HON", None),
                archivo_xml=str(fila.get("archivo_xml", "")).strip()
                if "archivo_xml" in df_actualizado.columns
                else None,
            )

        reporte = core.generar_reporte_texto(df_actualizado)
        ruta_reporte_dir = os.path.join(ruta_mes, "reporte_avance")
        os.makedirs(ruta_reporte_dir, exist_ok=True)
        ruta_reporte = os.path.join(
            ruta_reporte_dir,
            f"reporte_revision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
        )
        with open(ruta_reporte, "w", encoding="utf-8") as f:
            f.write(reporte)
        ui.log(f"Reporte: {ruta_reporte}", level="success")
        ui.emit("report.written", {"path": ruta_reporte})

        if ctx.supervised and not ui.confirm_yes_no(
            "Guardar Excel",
            "¿Guardar los cambios en Solicitud.xlsx?",
            default=True,
        ):
            ui.log("Excel no modificado (solo reporte).", level="warning")
            result = {"ok": True, "excel_saved": False, "report_path": ruta_reporte, **stats}
            ui.emit("session.summary", result)
            return result

        saved = core.guardar_excel(df_actualizado, ruta_excel, hoja, ui)
        result = {
            "ok": saved,
            "excel_saved": saved,
            "report_path": ruta_reporte,
            "log_file": ruta_log_file,
            **stats,
        }
        ui.emit("session.summary", result)
        if saved:
            ui.header("Validación finalizada", "Revisa el reporte en reporte_avance/")
        return result
