"""Etapa 4 — extracción datos XML al Excel."""
from __future__ import annotations

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import config
import schema_validator
import utils
from db import file_repository
from interaction.exceptions import SessionCancelled
from interaction.port import InteractionPort
from stages.context import Stage4Context
from stages.stage4 import extraction_core as core

RAIZ = core.RAIZ


class Stage4Service:
    def run(self, ctx: Stage4Context, ui: InteractionPort) -> dict:
        ui.header("Extracción de datos XML al Excel", "Rellena columnas *_XML en Solicitud")

        try:
            año, mes = utils.resolve_año_mes(RAIZ, ctx.year, ctx.month)
        except ValueError as e:
            ui.log(str(e), level="error")
            return {"ok": False}

        ruta_mes = os.path.join(RAIZ, año, mes)
        mes_num = config.MESES_ES.index(mes) + 1 if mes in config.MESES_ES else 0
        periodo_id = None
        if mes_num > 0:
            periodo_id = file_repository.get_or_create_periodo(
                anio=int(año),
                mes_num=mes_num,
                mes_nombre=mes,
            )

        ruta_logs = os.path.join(ruta_mes, "logs_extraccion_xml_excel")
        os.makedirs(ruta_logs, exist_ok=True)
        ruta_log_file = os.path.join(
            ruta_logs, datetime.now().strftime("extraccion_%Y%m%d_%H%M%S.log")
        )
        utils.configurar_logging(ruta_log_file)

        ruta_excel = os.path.join(ruta_mes, "Solicitud.xlsx")
        if not os.path.isfile(ruta_excel):
            ui.log(f"No se encontró Solicitud.xlsx en {ruta_mes}", level="error")
            return {"ok": False}

        ui.table(
            "Contexto",
            [
                ("Raíz", RAIZ),
                ("Período", f"{mes} {año}"),
                ("Carpeta mes", ruta_mes),
                ("Excel", ruta_excel),
                ("Logs", ruta_log_file),
            ],
        )

        if not ui.confirm_yes_no(
            "Continuar",
            "¿Extraer datos de los XML al Excel?",
            default=False,
        ):
            ui.log("Cancelado.", level="warning")
            return {"ok": False, "cancelled": True}

        wb = load_workbook(ruta_excel, read_only=True)
        hojas = wb.sheetnames
        wb.close()

        if ctx.sheet and ctx.sheet in hojas:
            hoja = ctx.sheet
        elif len(hojas) == 1:
            hoja = hojas[0]
        elif utils.is_non_interactive():
            hoja = utils.pick_excel_sheet(hojas)
        else:
            hoja = str(
                ui.choose_option(
                    "Hoja Excel",
                    "Seleccione la hoja donde agregar datos XML",
                    hojas,
                    icon="📄",
                )
            )

        df = pd.read_excel(ruta_excel, sheet_name=hoja, engine="openpyxl")

        con_xml = 0
        if "archivo_xml" in df.columns:
            con_xml = int(df["archivo_xml"].astype(str).str.strip().ne("").sum())
        ui.emit("scan.ready", {"rows_with_xml": con_xml, "total_rows": len(df), "sheet": hoja})

        canonical_errors, canonical_warnings = schema_validator.validate_for_stage(
            df, "stage4_extraccion_xml"
        )
        for w in canonical_warnings:
            logging.warning(f"[stage4] WARN {w}")
            ui.log(f"[schema] {w}", level="warning")
        for e in canonical_errors:
            logging.error(f"[stage4] ERROR {e}")
            ui.log(f"[schema] {e}", level="error")
        if canonical_errors and ctx.strict:
            ui.log("Validación estricta: abortando.", level="error")
            return {"ok": False}

        hay_datos_ok = (
            df["Observaciones_XML"].astype(str).str.strip().str.lower().eq("datos extraídos ok").any()
            if "Observaciones_XML" in df.columns
            else False
        )
        sobrescribir_ok = self._resolve_overwrite(ui, ctx, hay_datos_ok)

        if ctx.supervised and not ui.confirm_yes_no(
            "Procesar",
            f"¿Procesar {len(df)} filas de «{hoja}»?",
            default=False,
        ):
            return {"ok": False, "cancelled": True}

        try:
            df, stats = core.procesar_filas(
                df,
                ruta_mes,
                periodo_id=periodo_id,
                sobrescribir_ok=sobrescribir_ok,
                ui=ui,
            )
        except SessionCancelled:
            return {"ok": False, "cancelled": True}

        if ctx.supervised and not ui.confirm_yes_no(
            "Guardar Excel",
            "¿Guardar Solicitud.xlsx con los datos extraídos?",
            default=True,
        ):
            result = {"ok": True, "excel_saved": False, **stats}
            ui.emit("session.summary", result)
            return result

        saved = core.guardar_excel(df, ruta_excel, hoja, ui)
        result = {"ok": saved, "excel_saved": saved, "log_file": ruta_log_file, **stats}
        ui.emit("session.summary", result)
        if saved:
            ui.header("Extracción finalizada", f"Éxitos: {stats.get('exitos', 0)}")
        return result

    def _resolve_overwrite(
        self, ui: InteractionPort, ctx: Stage4Context, hay_datos_ok: bool
    ) -> bool:
        if not hay_datos_ok:
            return False
        if ctx.overwrite_ok is not None:
            return ctx.overwrite_ok
        if utils.is_non_interactive():
            ui.log("No se sobrescriben filas con «Datos extraídos OK».", level="info")
            return False
        return ui.confirm_yes_no(
            "Sobrescribir",
            "¿Sobrescribir filas que ya tienen «Datos extraídos OK»?",
            default=False,
        )
