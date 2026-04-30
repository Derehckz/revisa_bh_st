import os
import re
import logging
from datetime import datetime
import pandas as pd
from rich.progress import Progress, BarColumn, TimeElapsedColumn, TextColumn
import argparse
import tempfile
from concurrent.futures import ThreadPoolExecutor
from multiprocessing import cpu_count
import utils

# Opcionales para OCR/inspección
try:
    import fitz  # PyMuPDF
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False

try:
    # Para formato condicional y checkboxes (usamos colores)
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def buscar_rut_en_texto(texto):
    m = re.search(r"(\d{7,8}-[0-9kK])", texto)
    if m:
        return m.group(1)
    m2 = re.search(r"(\d{7,8})", texto)
    return m2.group(1) if m2 else None


def intentar_extraer_rut_de_archivos(ruta_docente):
    """Intentar localizar un RUT en los nombres de archivos dentro de la carpeta."""
    for nombre in os.listdir(ruta_docente):
        m = re.search(r"(\d{7,8}-[0-9kK])", nombre)
        if m:
            return m.group(1)
        m2 = re.search(r"(\d{7,8})", nombre)
        if m2:
            return m2.group(1)
    return None


def renombrar_seguro(src, dst):
    """Renombra `src` a `dst`. Si dst existe, añade sufijo incremental. Devuelve nuevo nombre o None en error."""
    base, ext = os.path.splitext(dst)
    destino = dst
    i = 1
    while os.path.exists(destino):
        destino = f"{base}_{i}{ext}"
        i += 1
    try:
        os.replace(src, destino)
        return os.path.basename(destino)
    except OSError as e:
        logging.warning(f"No se pudo renombrar {src} -> {destino}: {e}")
        return None


def estandarizar_nombres_en_carpeta(ruta_docente, rut, opts):
    """Renombra archivos dentro de una carpeta a los nombres estandarizados BHE_/CP_/IA_/CT_.
    Devuelve una lista de tuples (origen, destino) con las renombraciones realizadas.
    """
    cambios = []
    if not rut:
        return cambios

    # usar rut en formato limpio (con DV si tiene guion)
    rut_limpio = rut

    for nombre in os.listdir(ruta_docente):
        ruta = os.path.join(ruta_docente, nombre)
        if not os.path.isfile(ruta):
            continue
        low = nombre.lower()
        tipo = None
        if low.startswith('bhe_') or 'boleta' in low or 'honorarios' in low:
            tipo = 'BHE'
        elif low.startswith('ct_') or low.startswith('cont') or low.startswith('ct') or 'contrato' in low:
            tipo = 'CT'
        elif low.startswith('ia_') or 'informe' in low or 'actividades' in low:
            tipo = 'IA'
        elif low.startswith('cp_') or low.startswith('cp') or 'comprobante' in low:
            tipo = 'CP'

        if tipo:
            ext = os.path.splitext(nombre)[1] or '.pdf'
            nuevo = f"{tipo}_{rut_limpio}{ext}"
            destino = os.path.join(ruta_docente, nuevo)
            if opts.dry_run:
                cambios.append((nombre, nuevo))
            else:
                nuevo_nombre = renombrar_seguro(ruta, destino)
                if nuevo_nombre:
                    cambios.append((nombre, nuevo_nombre))

    return cambios

# Inicialización
console = utils.console

# Configuración
import config
RAIZ = config.RAIZ
ZONA_HORARIA = config.ZONA_HORARIA


def extraer_rut_y_nombre_de_carpeta(nombre_carpeta):
    """Intentar extraer RUT (formato 12345678-9) y nombre desde el nombre de la carpeta.
    Soporta formatos como 'EMPLID_Nombre' o '12345678-9_Nombre' o '12345678-9_Nombre, Apellidos'.
    """
    rut = ""
    nombre = nombre_carpeta
    # Buscar RUT con DV
    m = re.search(r"(\d{7,8}-[0-9kK])", nombre_carpeta)
    if m:
        rut = m.group(1)
        # intentar extraer nombre después del guion bajo
        parts = nombre_carpeta.split('_', 1)
        if len(parts) == 2:
            nombre = parts[1].replace('_', ' ').strip()
        else:
            nombre = nombre_carpeta
    else:
        # si no hay RUT, si existe '_' separar EMPLID y nombre
        parts = nombre_carpeta.split('_', 1)
        if len(parts) == 2:
            nombre = parts[1].replace('_', ' ').strip()
            # si primera parte parece un número, usar como rut sin dv
            if parts[0].isdigit():
                rut = parts[0]

    return rut, nombre


def analizar_carpeta_docente(ruta_docente, force_ocr=False):
    """Analiza los archivos dentro de la carpeta del docente y devuelve un dict con flags y observación.
    Observacion: lista de archivos presentes; si algún archivo fue identificado por OCR (porque su nombre no indicaba el tipo),
    se añade "(OCR: TIPOS)" junto al nombre del archivo.
    """
    archivos = [f for f in os.listdir(ruta_docente) if os.path.isfile(os.path.join(ruta_docente, f))]
    archivos_lower = [f.lower() for f in archivos]

    # Detección por nombre (rápida)
    tiene_ct = any(f.startswith('ct_') or f.startswith('cont') or f.startswith('ct') for f in archivos_lower)
    tiene_ia = any(f.startswith('ia_') or 'informe' in f for f in archivos_lower)
    # Boletas son siempre con prefijo BHE_
    tiene_bh = any(f.startswith('bhe_') or f.startswith('bhe-') or 'honorarios' in f for f in archivos_lower)
    tiene_cp = any(f.startswith('cp_') or f.startswith('cp') or 'comprobante' in f for f in archivos_lower)

    # Registrar detecciones por OCR por archivo cuando nombre no indicaba el tipo
    from collections import defaultdict
    ocr_map = defaultdict(set)

    # Si se solicita OCR y está disponible, intentar identificar tipos por contenido en PDFs
    if OCR_AVAILABLE and force_ocr and archivos:
        kw_bh = ['boleta', 'honorarios', 'boleta de honorarios']
        kw_ct = ['contrato', 'convenio', 'contrat']
        kw_ia = ['informe', 'actividades', 'informe de actividades']
        kw_cp = ['comprobante', 'pagos', 'pago', 'comprobante de pago']

        for f in archivos:
            if all([tiene_bh, tiene_ct, tiene_ia, tiene_cp]):
                break
            if not f.lower().endswith('.pdf'):
                continue
            ruta = os.path.join(ruta_docente, f)
            texto = ''
            try:
                import fitz
                with fitz.open(ruta) as doc:
                    for page in doc:
                        texto += page.get_text()
            except Exception:
                texto = ''
            if not texto:
                try:
                    images = convert_from_path(ruta, dpi=200)
                    for img in images:
                        texto += pytesseract.image_to_string(img, lang='spa') + ' '
                except Exception:
                    texto = ''

            txt = texto.lower()
            # Para cada tipo, si el archivo no mostraba por nombre el tipo, y OCR lo indica, marcarlo
            if any(k in txt for k in kw_bh):
                if not any(x in f.lower() for x in ('bhe_', 'honorarios', 'boleta')):
                    ocr_map[f].add('BHE')
                tiene_bh = True
            if any(k in txt for k in kw_ct):
                if not any(x in f.lower() for x in ('ct_', 'cont', 'contrato')):
                    ocr_map[f].add('CT')
                tiene_ct = True
            if any(k in txt for k in kw_ia):
                if not any(x in f.lower() for x in ('ia_', 'informe', 'actividades')):
                    ocr_map[f].add('IA')
                tiene_ia = True
            if any(k in txt for k in kw_cp):
                if not any(x in f.lower() for x in ('cp_', 'comprobante')):
                    ocr_map[f].add('CP')
                tiene_cp = True

    # Construir Observacion: lista de archivos y marcar OCR si corresponde
    obs_items = []
    for f in archivos:
        if f in ocr_map and ocr_map[f]:
            tipos = ','.join(sorted(ocr_map[f]))
            obs_items.append(f"{f} (OCR: {tipos})")
        else:
            obs_items.append(f)

    observacion = '; '.join(obs_items)

    return {
        'CT': tiene_ct,
        'IA': tiene_ia,
        'BH': tiene_bh,
        'CP': tiene_cp,
        'Observacion': observacion,
        'Archivos': archivos
    }


def _estado_desde_cantidad(n):
    """Mapea cantidad de archivos (0-4) a estado: Completo, Revisar, Incompleto."""
    if n >= 4:
        return 'Completo'
    if n == 3:
        return 'Revisar'
    return 'Incompleto'


def _calcular_estadisticas_revision(df):
    """Calcula totales y porcentajes para el resumen de revisión. df debe tener columnas CT, IA, BH, CP."""
    total_carpetas = len(df)
    df = df.copy()
    df['present_count'] = df[['CT', 'IA', 'BH', 'CP']].sum(axis=1)
    df['Estado'] = df['present_count'].apply(_estado_desde_cantidad)
    total_esperado = total_carpetas * 4
    total_encontrado = int(df['present_count'].sum())
    total_faltante = total_esperado - total_encontrado
    pct_presente = (total_encontrado / total_esperado * 100) if total_esperado else 0.0
    carpetas_completas = int((df['present_count'] == 4).sum())
    return {
        'df': df,
        'total_carpetas': total_carpetas,
        'total_esperado': total_esperado,
        'total_encontrado': total_encontrado,
        'total_faltante': total_faltante,
        'pct_presente': pct_presente,
        'carpetas_completas': carpetas_completas,
    }


def _guardar_excel_revision(ruta_mes, df, estadisticas):
    """Escribe el Excel de revisión en ruta_mes y aplica formato con openpyxl si está disponible.
    Devuelve True si se guardó correctamente, False en caso contrario.
    """
    salida_excel = os.path.join(ruta_mes, 'revision_carpetas.xlsx')
    total_carpetas = estadisticas['total_carpetas']
    total_esperado = estadisticas['total_esperado']
    total_encontrado = estadisticas['total_encontrado']
    total_faltante = estadisticas['total_faltante']
    pct_presente = estadisticas['pct_presente']
    carpetas_completas = estadisticas['carpetas_completas']

    with pd.ExcelWriter(salida_excel, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name='Revisión')
        resumen = pd.DataFrame([{
            'Total Carpetas': total_carpetas,
            'Total Esperado (4 por carpeta)': total_esperado,
            'Total Encontrado': total_encontrado,
            'Total Faltante': total_faltante,
            'Porcentaje Presente': f"{pct_presente:.2f}%",
            'Carpetas Completas (4/4)': carpetas_completas
        }])
        resumen.to_excel(writer, index=False, sheet_name='Resumen')

    if OPENPYXL_AVAILABLE:
        _aplicar_formato_condicional_excel(salida_excel, ruta_mes)

    utils.print_success(f"Archivo de revisión guardado en: {salida_excel}")
    utils.print_info(
        f"📊 {total_encontrado}/{total_esperado} archivos presentes ({pct_presente:.2f}%). "
        f"Carpetas completas: {carpetas_completas}/{total_carpetas}. Faltan: {total_faltante} archivos."
    )
    return True


def _aplicar_formato_condicional_excel(salida_excel, ruta_mes):
    """Aplica formato condicional (colores, anchos, hipervínculos) al Excel de revisión."""
    from openpyxl.styles import Font, Alignment
    wb = load_workbook(salida_excel)
    ws = wb['Revisión']
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    header = [c.value for c in ws[1]]
    for col_cell in ws[1]:
        col_cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    col_idx_map = {name: (header.index(name) + 1) for name in header}

    for col_name in ('CT', 'IA', 'BH', 'CP'):
        if col_name in col_idx_map:
            col_idx = col_idx_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                truthy = False
                if isinstance(val, bool):
                    truthy = val
                elif isinstance(val, (int, float)):
                    truthy = val != 0
                elif isinstance(val, str):
                    truthy = val.strip().lower() in ('si', 'sí', 'true', '1', '✅ sí', '✅')
                cell.value = 'Sí' if truthy else 'No'
                cell.alignment = Alignment(horizontal='center')
                cell.fill = green if truthy else red

    if 'Observacion' in col_idx_map:
        obs_idx = col_idx_map['Observacion']
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=obs_idx)
            cell.alignment = Alignment(wrap_text=True)

    if 'Ubicacion Carpeta' in col_idx_map:
        ub_idx = col_idx_map['Ubicacion Carpeta']
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ub_idx)
            rel = cell.value
            if rel:
                target = os.path.join(ruta_mes, str(rel))
                if os.path.exists(target):
                    cell.hyperlink = target
                    cell.style = 'Hyperlink'

    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
            except (TypeError, ValueError):
                val = ''
            if len(val) > max_len:
                max_len = len(val)
        adjusted_width = min(max(10, max_len + 2), 60)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = adjusted_width

    if 'Estado' in col_idx_map:
        est_idx = col_idx_map['Estado']
        yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=est_idx)
            val = (cell.value or '').strip()
            if val == 'Completo':
                cell.fill = green
            elif val == 'Revisar':
                cell.fill = yellow
            else:
                cell.fill = red

    wb.save(salida_excel)


def eliminar_marcadores(ruta_mes, institutos):
    """Elimina todos los archivos '.revisado' bajo las carpetas de las instituciones del mes.
    Devuelve el número de marcadores eliminados."""
    count = 0
    for inst in institutos:
        base = os.path.join(ruta_mes, inst)
        for dp, ds, files in os.walk(base):
            if '.revisado' in files:
                p = os.path.join(dp, '.revisado')
                try:
                    os.remove(p)
                    count += 1
                except Exception:
                    logging.warning(f"No se pudo eliminar marcador {p}")
    return count


def ejecutar_trabajos(trabajos, procesos):
    """Ejecuta la lista de trabajos en paralelo y devuelve la lista de registros resultantes."""
    registros_local = []
    with ThreadPoolExecutor(max_workers=procesos) as executor:
        for res in executor.map(_worker_tuple, trabajos):
            if res:
                registros_local.append(res)
    return registros_local


def _worker_tuple(args_tuple):
    # helper to unpack for Pool
    institucion, ruta_mes, carpeta_docente, opts = args_tuple
    ruta_docente = os.path.join(ruta_mes, institucion, carpeta_docente)

    # Si existe marcador .revisado y no forzamos, comprobar si la carpeta está completa.
    # Si tiene los 4 archivos (CT, IA, BH, CP) saltar; si falta alguno, re-evaluar cada ejecución.
    marcado_path = os.path.join(ruta_docente, '.revisado')
    if os.path.exists(marcado_path) and not opts.force:
        try:
            info_check = analizar_carpeta_docente(ruta_docente, force_ocr=getattr(opts, 'ocr', False))
            if info_check.get('CT') and info_check.get('IA') and info_check.get('BH') and info_check.get('CP'):
                return None
            # else: continuar con el procesamiento porque falta algún comprobante
        except Exception:
            # si falla la comprobación, seguir y re-evaluar para mayor seguridad
            pass
    rut, nombre = extraer_rut_y_nombre_de_carpeta(carpeta_docente)

    # Si se pidió renombrar, intentar estandarizar nombres usando RUT extraído de carpeta o OCR
    if getattr(opts, 'rename', False):
        rut_candidate = rut or intentar_extraer_rut_de_archivos(ruta_docente)
        if not rut_candidate and OCR_AVAILABLE and getattr(opts, 'ocr', False):
            # intentar OCR en primeros PDFs para extraer RUT
            for f in os.listdir(ruta_docente):
                if not f.lower().endswith('.pdf'):
                    continue
                ruta_f = os.path.join(ruta_docente, f)
                try:
                    import fitz
                    texto = ''
                    with fitz.open(ruta_f) as doc:
                        for p in doc:
                            texto += p.get_text()
                except Exception:
                    texto = ''
                if not texto and OCR_AVAILABLE:
                    try:
                        images = convert_from_path(ruta_f, dpi=200)
                        for img in images:
                            texto += pytesseract.image_to_string(img, lang='spa') + ' '
                    except Exception:
                        texto = ''
                if texto:
                    found = buscar_rut_en_texto(texto)
                    if found:
                        rut_candidate = found
                        break

        if rut_candidate:
            # realizar renombrado antes del análisis para que Observacion refleje nombres finales
            _ = estandarizar_nombres_en_carpeta(ruta_docente, rut_candidate, opts)

    # Analizar después de (posible) renombrado; permitir forzar OCR con la flag
    info_files = analizar_carpeta_docente(ruta_docente, force_ocr=getattr(opts, 'ocr', False))

    # Si no dry-run y se pidió marcar, crear marcador
    if opts.mark and not opts.dry_run:
        try:
            with open(marcado_path, 'w', encoding='utf-8') as fh:
                fh.write(f"revisado: {datetime.now().isoformat()}\n")
        except Exception:
            pass

    return {
        'Institucion': institucion,
        'RUT': rut,
        'Nombre Docente': nombre,
        'Ubicacion Carpeta': os.path.join(institucion, carpeta_docente),
        'CT': info_files['CT'],
        'IA': info_files['IA'],
        'BH': info_files['BH'],
        'CP': info_files['CP'],
        'Observacion': info_files['Observacion']
    }


def main():
    parser = argparse.ArgumentParser(description='Revisa carpetas IP/CFT por mes y genera un Excel de revisión')
    parser.add_argument('--year', '-y', help='Año (carpeta)', type=str)
    parser.add_argument('--month', '-m', help='Mes (carpeta)', type=str)
    parser.add_argument('--institucion', '-i', help='Filtrar por institucion (IP o CFT)', type=str)
    parser.add_argument('--dry-run', action='store_true', help='No escribe marcadores ni archivos')
    parser.add_argument('--force', action='store_true', help='Forzar re-evaluación aunque exista .revisado')
    parser.add_argument('--no-mark', dest='mark', action='store_false', help='No crear archivo .revisado')
    parser.add_argument('--ocr', action='store_true', help='Forzar uso de OCR (si está disponible)')
    parser.add_argument('--rename', action='store_true', help='Estandarizar nombres de archivos (BHE_/CP_/IA_/CT_)')
    parser.add_argument('--processes', type=int, default=max(1, cpu_count()-1), help='Número de procesos paralelos')
    args = parser.parse_args()

    utils.print_header("🔎 REVISIÓN DE CARPETAS IP/CFT")

    # Selección año/mes
    años = utils.listar_carpetas(RAIZ)
    if not años:
        utils.print_error("No hay carpetas de año en la ruta configurada.")
        return

    año = args.year if args.year else utils.seleccionar_opcion(sorted(años), "Seleccione el año:", "🗓️")
    ruta_año = os.path.join(RAIZ, año)

    meses = [d for d in os.listdir(ruta_año) if os.path.isdir(os.path.join(ruta_año, d))]
    if not meses:
        utils.print_error(f"No hay carpetas de mes en {ruta_año}")
        return

    mes = args.month if args.month else utils.seleccionar_opcion(sorted(meses), "Seleccione el mes:", "🗓️")
    ruta_mes = os.path.join(ruta_año, mes)

    # Buscar carpetas IP y CFT dentro de la carpeta del mes
    institutos = []
    for d in os.listdir(ruta_mes):
        ruta_d = os.path.join(ruta_mes, d)
        if os.path.isdir(ruta_d) and d.upper() in ("IP", "CFT"):
            institutos.append(d)

    if not institutos:
        utils.print_warning("No se encontraron carpetas 'IP' o 'CFT' en el mes seleccionado.")
        return

    if args.institucion:
        institutos = [i for i in institutos if i.upper() == args.institucion.upper()]

    utils.print_info(f"Se analizarán las siguientes instituciones: {', '.join(institutos)}")

    # Preparar lista de trabajos
    trabajos = []
    for institucion in institutos:
        ruta_institucion = os.path.join(ruta_mes, institucion)
        docentes = [d for d in os.listdir(ruta_institucion) if os.path.isdir(os.path.join(ruta_institucion, d))]
        for carpeta_docente in docentes:
            trabajos.append((institucion, ruta_mes, carpeta_docente, args))

    # Ejecutar en paralelo
    procesos = max(1, args.processes)
    utils.print_info(f"Usando {procesos} procesos para analizar {len(trabajos)} carpetas...")
    registros = ejecutar_trabajos(trabajos, procesos)

    # Si no se obtuvo ningún registro, ofrecer opciones interactivas para re-evaluar
    if not registros:
        utils.print_warning("No se encontraron docentes para analizar. Puede deberse a carpetas ya marcadas como revisadas (.revisado).")
        opciones = [
            "Re-evaluar ahora (ignorar marcadores .revisado)",
            "Eliminar marcadores .revisado y reintentar",
            "Salir"
        ]
        eleccion = utils.seleccionar_opcion(opciones, "¿Qué quieres hacer?", "❓")
        if eleccion == opciones[0]:
            args.force = True
            utils.print_info("Re-evaluando ignorando marcadores (.revisado)...")
            registros = ejecutar_trabajos(trabajos, procesos)
        elif eleccion == opciones[1]:
            if args.dry_run:
                utils.print_warning("Estás en modo --dry-run: no se eliminarán los marcadores. Ejecuta sin --dry-run para borrar.")
            else:
                n = eliminar_marcadores(ruta_mes, institutos)
                utils.print_info(f"Se eliminaron {n} marcadores. Reintentando análisis...")
                registros = ejecutar_trabajos(trabajos, procesos)
        else:
            utils.print_warning("Saliendo sin procesar.")
            return

    # Guardar resultados en Excel dentro de la carpeta del mes
    df = pd.DataFrame(registros)
    if df.empty:
        utils.print_warning("No se encontraron docentes para analizar.")
        return

    estadisticas = _calcular_estadisticas_revision(df)
    df = estadisticas['df']

    try:
        _guardar_excel_revision(ruta_mes, df, estadisticas)
    except (OSError, IOError, PermissionError) as e:
        console.print(Panel.fit(f"❌ Error guardando Excel: {e}", style="bold red"))


if __name__ == '__main__':
    main()
