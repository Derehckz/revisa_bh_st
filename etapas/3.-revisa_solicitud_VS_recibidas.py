#!/usr/bin/env python3
import _sys_path  # noqa: E402
import os
import re
import sys
import argparse
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import logging
from datetime import datetime
from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn
from rich.table import Table
from rich.panel import Panel
from rich.text import Text
from rich.logging import RichHandler
import utils
import schema_validator
from db import boleta_repository, docente_repository, file_repository
from db.key_builder import build_boleta_key

console = utils.console

# Configuración global
import config
RAIZ = config.RAIZ
PREFIJO = config.PREFIJO

# Funciones para Rich UI

def mostrar_progreso_rich(total):
    """Inicia una barra de progreso Rich y devuelve (task_id, progress_obj)"""
    progress = Progress(
        TextColumn("[cyan]🔄 Procesando:"),
        BarColumn(bar_width=None),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TimeElapsedColumn(),
        console=console,
    )
    task = progress.add_task("", total=total)
    progress.start()
    return task, progress

def generar_reporte_texto(df):
    from datetime import datetime

    total = len(df)
    recibidos = df['Estado_Recepcion'].fillna('').astype(str).str.strip().str.upper().isin(['RECIBIDO', 'RECIBIDO CON ERROR']).sum()
    porcentaje = (recibidos / total * 100) if total > 0 else 0

    resumen = df.groupby(['Estado_Recepcion', 'Observaciones']).size().reset_index(name='Cantidad')

    def fit_text(texto, ancho):
        texto = str(texto)
        if len(texto) > ancho:
            return texto[:ancho-3] + "..."
        return texto.ljust(ancho)

    lineas = []
    lineas.append("╔" + "═"*100 + "╗")
    lineas.append("║" + fit_text("Reporte Detallado de Validación de Recepción", 100).center(100) + "║")
    lineas.append("║" + fit_text(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 100).center(100) + "║")
    lineas.append("╚" + "═"*100 + "╝")

    lineas.append(f"Total registros analizados: {total}")
    lineas.append(f"Total recibidos (OK + con error): {recibidos} ({porcentaje:.2f}%)")
    lineas.append("─"*104)

    # Resumen agrupado
    lineas.append(f"{fit_text('Estado',20)} │ {fit_text('Observación',50)} │ {'Cantidad':>8}")
    lineas.append("─"*104)
    for _, fila in resumen.iterrows():
        estado = fila['Estado_Recepcion']
        obs = fila['Observaciones']
        cant = fila['Cantidad']
        lineas.append(f"{fit_text(estado,20)} │ {fit_text(obs,50)} │ {str(cant).rjust(8)}")
    lineas.append("─"*104)

    # Separar OK y errores
    df_ok = df[df['Estado_Recepcion'].fillna('').astype(str).str.strip().str.upper() == 'RECIBIDO']
    df_err = df[df['Estado_Recepcion'].fillna('').astype(str).str.strip().str.upper() != 'RECIBIDO']


    # Definición columnas comunes para ambos listados
    headers = [
        ('RUT', 15),
        ('Nombre Docente', 25),
        ('Nombre / Razón', 30),
        ('Estado', 15),
        ('Observaciones', 50),
    ]

    def header_line():
        return " │ ".join(fit_text(name, width) for name, width in headers)

    def separator_line():
        return "─" * (sum(width for _, width in headers) + 3 * (len(headers) -1))

    # Listado OK
    lineas.append("\nRegistros OK (RECIBIDO):")
    lineas.append(header_line())
    lineas.append(separator_line())
    for _, fila in df_ok.iterrows():
        rut = fila.get('RUT_SIN_DV', '')
        nombre_docente = fila.get('NAME', '')
        nombre_razon = fila.get('RUT RAZON', '')
        estado = fila.get('Estado_Recepcion', 'RECIBIDO')
        obs = fila.get('Observaciones', 'OK')
        lineas.append(
            " │ ".join([
                fit_text(rut, 15),
                fit_text(nombre_docente, 25),
                fit_text(nombre_razon, 30),
                fit_text(estado, 15),
                fit_text(obs, 50)
            ])
        )

    # Listado Pendientes / Errores
    lineas.append("\nRegistros Pendientes / Con Error:")
    lineas.append(header_line())
    lineas.append(separator_line())
    for _, fila in df_err.iterrows():
        rut = fila.get('RUT_SIN_DV', '')
        nombre_docente = fila.get('NAME', '')
        nombre_razon = fila.get('RUT RAZON', '')
        estado = fila.get('Estado_Recepcion', '')
        obs = fila.get('Observaciones', '')
        lineas.append(
            " │ ".join([
                fit_text(rut, 15),
                fit_text(nombre_docente, 25),
                fit_text(nombre_razon, 30),
                fit_text(estado, 15),
                fit_text(obs, 50)
            ])
        )

    return "\n".join(lineas)



def imprimir_resumen_rich(df, revisados_inicio, revisados_final):
    """Imprime un resumen final en consola usando tablas y paneles Rich"""
    resumen = df.groupby(['Estado_Recepcion', 'Observaciones']).size().reset_index(name='Cantidad')
    table = Table(show_lines=True)
    table.add_column("Estado", justify="center", style="bold")
    table.add_column("Observación", style="dim")
    table.add_column("Cantidad", justify="right")

    for _, fila in resumen.iterrows():
        estado = fila['Estado_Recepcion']
        observacion = fila['Observaciones']
        cantidad = str(fila['Cantidad'])
        style = (
            "green" if estado == "RECIBIDO" else
            "yellow" if estado == "RECIBIDO CON ERROR" else
            "red"
        )
        table.add_row(estado, observacion, cantidad, style=style)

    panel = Panel(table,
                  title=f"Revisados inicio: {revisados_inicio} • final: {revisados_final}",
                  border_style="magenta")
    console.print(panel)

def mostrar_porcentaje_recepcion(df, total):
    recibidos = df['Estado_Recepcion'].fillna('').astype(str).str.strip().str.upper().isin(['RECIBIDO', 'RECIBIDO CON ERROR']).sum()
    porcentaje = (recibidos / total) * 100 if total > 0 else 0

    texto = Text()
    texto.append(f"Recibidos: {recibidos} / {total}  ", style="bold green")
    texto.append(f"({porcentaje:.2f}%)", style="bold yellow")

    panel = Panel(
        texto,
        title="📊 Progreso de Recepción",
        border_style="bright_blue",
        expand=False
    )
    console.print(panel)


# Funciones originales de extracción y comparación

def extraer_datos_xml(ruta_xml):
    try:
        tree = ET.parse(ruta_xml)
        root = tree.getroot()

        rutEmisor_elem = utils.find_element_ignore_ns(root, 'rutEmisor')
        rutReceptor_elem = utils.find_element_ignore_ns(root, 'rutReceptor')
        dvReceptor_elem = utils.find_element_ignore_ns(root, 'dvReceptor')
        totalHonorarios_elem = utils.find_element_ignore_ns(root, 'totalHonorarios')
        descripcionLinea_elem = utils.find_element_ignore_ns(root, 'descripcionLinea')

        th_raw = totalHonorarios_elem.text.strip() if totalHonorarios_elem is not None and totalHonorarios_elem.text else ""
        try:
            total_h = float(th_raw) if th_raw not in ("", None) else None
        except (TypeError, ValueError):
            total_h = None
        return {
            'rutEmisor': rutEmisor_elem.text.strip() if rutEmisor_elem is not None else '',
            'rutReceptor': rutReceptor_elem.text.strip() if rutReceptor_elem is not None else '',
            'dvReceptor': dvReceptor_elem.text.strip() if dvReceptor_elem is not None else '',
            'totalHonorarios': total_h,
            'descripcionLinea': descripcionLinea_elem.text.strip() if descripcionLinea_elem is not None else '',
        }
    except (ET.ParseError, OSError, ValueError, TypeError) as e:
        return {'error': f'Error leyendo XML: {e}'}


def normalizar_rut_con_dv(rut):
    # Delegar a utils para normalización consistente
    return utils.normalizar_rut_con_dv(rut)


def separar_rut_dv(rut_completo):
    rut = normalizar_rut_con_dv(rut_completo)
    return rut[:-1], rut[-1:] if len(rut) > 1 else ('', '')


def _rut_variantes_prefijo_archivo(rut_cualquier_formato: str) -> list[str]:
    """Cuerpo numérico del RUT y variante sin ceros a la izquierda (nombres bhe_... en disco)."""
    d = utils.normalizar_rut_digits(rut_cualquier_formato)
    if not d:
        return []
    stripped = d.lstrip("0") or "0"
    return list(dict.fromkeys([d, stripped] if stripped != d else [d]))


def archivos_relacionados(ruta_carpeta, rut_normalizado):
    archivos = os.listdir(ruta_carpeta)
    alts = "|".join(re.escape(v) for v in _rut_variantes_prefijo_archivo(rut_normalizado))
    if not alts:
        alts = re.escape(str(rut_normalizado))
    patron = re.compile(rf"^{PREFIJO}(?:{alts})[-_\d]*\.(pdf|xml)$", re.IGNORECASE)
    pdfs, xmls = [], []
    for f in archivos:
        if patron.match(f):
            (pdfs if f.lower().endswith('.pdf') else xmls).append(f)
    return pdfs, xmls


def comparar_datos(excel_fila, datos_xml):
    obs = []
    rut_sin_dv = normalizar_rut_con_dv(excel_fila.get('RUT_SIN_DV'))
    if normalizar_rut_con_dv(datos_xml.get('rutEmisor')) != rut_sin_dv:
        obs.append(f"rutEmisor XML ({datos_xml.get('rutEmisor')}) distinto a RUT_SIN_DV Excel ({excel_fila.get('RUT_SIN_DV')})")

    rut_razon = normalizar_rut_con_dv(excel_fila.get('RUT RAZON'))
    rut_receptor = normalizar_rut_con_dv(datos_xml.get('rutReceptor','') + datos_xml.get('dvReceptor',''))
    if separar_rut_dv(rut_receptor)[0] != separar_rut_dv(rut_razon)[0]:
        obs.append(f"rutReceptor XML ({rut_receptor}) distinto a RUT RAZON Excel ({excel_fila.get('RUT RAZON')})")

    try:
        monto_excel = float(excel_fila.get('CUS_TOT_HON', 0))
    except (TypeError, ValueError):
        monto_excel = 0
    monto_xml = datos_xml.get('totalHonorarios')
    if monto_xml is None or abs(monto_excel - monto_xml) > 1:
        obs.append(f"totalHonorarios XML ({monto_xml}) distinto a CUS_TOT_HON Excel ({monto_excel})")

    return obs

def cargar_xmls_por_rut(ruta_carpeta):
    xmls = [f for f in os.listdir(ruta_carpeta) if f.lower().endswith('.xml') and f.startswith(PREFIJO)]
    dct = {}
    for archivo in xmls:
        datos = extraer_datos_xml(os.path.join(ruta_carpeta, archivo))
        if 'error' in datos: continue
        key = (normalizar_rut_con_dv(datos['rutEmisor']), normalizar_rut_con_dv(datos['rutReceptor']+datos['dvReceptor']))
        dct[key] = datos
    return dct

def obtener_pdfs_con_monto_esperado(ruta_carpeta, rut_normalizado, monto_esperado, xml_usados=None, margen=1):
    archivos = os.listdir(ruta_carpeta)
    pdfs_validos = []
    for f in archivos:
        if not f.lower().endswith('.pdf'):
            continue
        # Solo archivos que empiecen con el prefijo + rut
        lowf = f.lower()
        if not any(lowf.startswith((PREFIJO + v).lower()) for v in _rut_variantes_prefijo_archivo(rut_normalizado)):
            continue
        # Sufijo tras el RUT: tomar parte tras última coincidencia de prefijo+rut
        sufijo_part = lowf
        for v in _rut_variantes_prefijo_archivo(rut_normalizado):
            pref = (PREFIJO + v).lower()
            if lowf.startswith(pref):
                sufijo_part = lowf[len(pref) :]
                break
        sufijo = sufijo_part.split(".")[0].strip("-_")
        if not sufijo.isdigit():
            continue
        base_xml = f[:-4] + '.xml'
        ruta_xml = os.path.join(ruta_carpeta, base_xml)
        if not os.path.isfile(ruta_xml):
            continue
        if xml_usados and base_xml in xml_usados:
            continue
        datos_xml = extraer_datos_xml(ruta_xml)
        if 'error' in datos_xml:
            continue
        monto_xml = datos_xml.get('totalHonorarios')
        if monto_xml is not None and abs(monto_xml - monto_esperado) <= margen:
            pdfs_validos.append(f)
    return pdfs_validos

def extraer_sufijo_archivo(nombre_archivo):
    # Extrae el sufijo numérico del archivo, ej: "bhe_6717063-118.pdf" / "bhe_6717063_118.xml"
    base = os.path.splitext(nombre_archivo)[0]
    m = re.search(r"[-_](\d+)$", base)
    if m:
        return m.group(1)
    return None

def es_provisionado(texto):
    texto = str(texto).lower()
    patrones = ["provisionado", "provisonado", "provs"]
    return any(p in texto for p in patrones)

def procesar_filas(df, ruta_carpeta):
    df = df.copy()

    columnas_str = ['Estado_Recepcion', 'Observaciones', 'archivo_xml']
    for col in columnas_str:
        if col not in df.columns:
            df[col] = ''
        df[col] = df[col].fillna('').astype(str)

    xml_usados = set()
    archivos = os.listdir(ruta_carpeta)
    patron_xml = re.compile(rf"^{PREFIJO}.*\.xml$", re.IGNORECASE)
    xml_todos = [f for f in archivos if patron_xml.match(f)]

    total = len(df)
    revis_ini = df['Estado_Recepcion'].fillna('').astype(str).str.strip().str.upper().isin(['RECIBIDO','RECIBIDO CON ERROR']).sum()
    task, progress = mostrar_progreso_rich(total)

    df['__provisionado__'] = df['GLOSA'].astype(str).str.lower().str.contains("provisionado")
    df.sort_values(by=['RUT_SIN_DV', 'CUS_TOT_HON', '__provisionado__'], ascending=[True, True, False], inplace=True)

    for idx, fila in df.iterrows():
        progress.advance(task)

        if pd.isna(fila.get('RUT_SIN_DV')) or pd.isna(fila.get('RUT RAZON')):
            df.at[idx,'Estado_Recepcion'] = 'NO RECIBIDO'
            df.at[idx,'Observaciones'] = 'RUT vacío'
            df.at[idx,'archivo_xml'] = ''
            logging.warning(f"Fila {idx+1}: RUT vacío. Estado: NO RECIBIDO.")
            continue

        rut_sin_dv = normalizar_rut_con_dv(fila['RUT_SIN_DV'])
        rut_razon = normalizar_rut_con_dv(fila['RUT RAZON'])
        monto_excel = float(fila.get('CUS_TOT_HON', 0))
        variants_pref = _rut_variantes_prefijo_archivo(fila['RUT_SIN_DV'])
        if 'EMPLID' in df.columns and fila.get('EMPLID') is not None and not pd.isna(fila.get('EMPLID')):
            variants_pref = list(
                dict.fromkeys(variants_pref + _rut_variantes_prefijo_archivo(str(fila.get('EMPLID'))))
            )

        datos_validos = None
        archivo_xml_valido = None
        xml_existente_con_fallo_institucion = False

        # Buscar XML válido primero
        for archivo_xml in xml_todos:
            low_xml = archivo_xml.lower()
            if not any(low_xml.startswith(f"{PREFIJO}{v}") for v in variants_pref):
                continue
            if archivo_xml in xml_usados:
                continue

            sufijo_xml = extraer_sufijo_archivo(archivo_xml)
            if sufijo_xml is None:
                continue

            datos_xml = extraer_datos_xml(os.path.join(ruta_carpeta, archivo_xml))
            if 'error' in datos_xml:
                continue

            monto_xml = datos_xml.get('totalHonorarios')
            if monto_xml is None or abs(monto_excel - monto_xml) > 1:
                continue

            rut_emisor_xml = normalizar_rut_con_dv(datos_xml.get('rutEmisor'))
            rut_receptor_xml = normalizar_rut_con_dv(datos_xml.get('rutReceptor','') + datos_xml.get('dvReceptor',''))

            if rut_emisor_xml != rut_sin_dv:
                xml_existente_con_fallo_institucion = True
                continue

            if separar_rut_dv(rut_receptor_xml)[0] != separar_rut_dv(rut_razon)[0]:
                continue

            nombre_pdf_esperado = archivo_xml[:-4] + '.pdf'
            if not os.path.isfile(os.path.join(ruta_carpeta, nombre_pdf_esperado)):
                continue

            descripcion_xml = datos_xml.get('descripcionLinea', '').lower()
            glosa_excel = str(fila.get('GLOSA', '')).lower()

            if es_provisionado(glosa_excel) != es_provisionado(descripcion_xml):
                continue

            datos_validos = datos_xml
            archivo_xml_valido = archivo_xml
            break

        # Si no hay XML válido, revisar PDF sin XML asociado
        if datos_validos is None:
            estado = ""
            obs_text = ""

            if xml_existente_con_fallo_institucion:
                estado = "RECIBIDO CON ERROR"
                obs_text = "XML existe pero institución de emisión incorrecta o inconsistente"
                logging.error(f"Fila {idx+1}: {obs_text}.")
            else:
                pdf_sin_xml = False
                bases_con_xml = {os.path.splitext(xml_name)[0] for xml_name in archivos if xml_name.lower().endswith(".xml") and xml_name.lower().startswith(PREFIJO)}

                for f in archivos:
                    low_pdf = f.lower()
                    if not low_pdf.endswith('.pdf'):
                        continue
                    if not any(low_pdf.startswith(f"{PREFIJO}{v}") for v in variants_pref):
                        continue
                    base_pdf = os.path.splitext(f)[0]
                    if base_pdf in bases_con_xml:
                        continue

                    pdf_sin_xml = True
                    break

                if pdf_sin_xml:
                    estado = "RECIBIDO CON ERROR"
                    obs_text = "PDF existe para monto esperado, pero no tiene XML asociado"
                    logging.warning(f"Fila {idx+1}: {obs_text}.")
                else:
                    estado = "NO RECIBIDO"
                    obs_text = "No se encontró PDF ni XML válido para el monto esperado"
                    logging.error(f"Fila {idx+1}: {obs_text}.")

            df.at[idx,'Estado_Recepcion'] = estado
            df.at[idx,'Observaciones'] = obs_text
            df.at[idx,'archivo_xml'] = ''
            continue

        # Si hay XML válido
        xml_usados.add(archivo_xml_valido)
        obs = comparar_datos(fila, datos_validos)
        if not obs:
            df.at[idx,'Estado_Recepcion'] = 'RECIBIDO'
            df.at[idx,'archivo_xml'] = archivo_xml_valido
            if es_provisionado(fila.get('GLOSA', '')):
                df.at[idx,'Observaciones'] = 'OK; OJO ES PROVISIONADO'
            else:
                df.at[idx,'Observaciones'] = 'OK'
            logging.info(f"Fila {idx+1}: Datos correctos. Estado: RECIBIDO.")
        else:
            df.at[idx,'Estado_Recepcion'] = 'RECIBIDO CON ERROR'
            df.at[idx,'archivo_xml'] = archivo_xml_valido
            df.at[idx,'Observaciones'] = '; '.join(obs)
            logging.warning(f"Fila {idx+1}: Observaciones: {'; '.join(obs)}")

    df.drop(columns='__provisionado__', inplace=True)
    progress.stop()
    revis_fin = df['Estado_Recepcion'].str.strip().str.upper().isin(['RECIBIDO','RECIBIDO CON ERROR']).sum()

    imprimir_resumen_rich(df, revis_ini, revis_fin)
    logging.info(f"Resumen final - Revisados inicio: {revis_ini}, final: {revis_fin}, Total: {total}")
    utils.print_success("Proceso completado con éxito.")

    return df


def guardar_excel(df, ruta_excel, hoja):
    try:
        hojas = pd.read_excel(ruta_excel, sheet_name=None, engine='openpyxl')
        hojas[hoja] = df

        def _writer(tmp_path):
            with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='w') as writer:
                for nombre_hoja, df_hoja in hojas.items():
                    df_hoja.to_excel(writer, index=False, sheet_name=nombre_hoja)

        utils.atomic_excel_write(ruta_excel, _writer)
        utils.print_success(f"Archivo guardado correctamente: {ruta_excel}")
    except (OSError, IOError) as e:
        utils.print_error(f"Error guardando archivo Excel: {e}")


def seleccionar_opcion(lista, mensaje, icono=""):
    # Delegar a utils para mantener una única implementación compartida
    return utils.seleccionar_opcion(lista, mensaje, icono)


def main(args=None):
    if args is None:
        args = argparse.Namespace(strict=False, yes=False, year=None, month=None)
    utils.apply_non_interactive_from_args(args)
    utils.print_header("📂 VALIDACIÓN DE RECEPCIÓN PDF Y XML")
    try:
        año, mes = utils.resolve_año_mes(RAIZ, getattr(args, "year", None), getattr(args, "month", None))
    except ValueError as e:
        utils.print_error(str(e))
        return
    ruta_año = os.path.join(RAIZ, año)
    ruta_mes = os.path.join(ruta_año, mes)
    mes_num = config.MESES_ES.index(mes) + 1 if mes in config.MESES_ES else 0

    ruta_logs = os.path.join(ruta_mes,"logs_revision")
    os.makedirs(ruta_logs, exist_ok=True)
    ruta_log_file = os.path.join(ruta_logs, datetime.now().strftime("revision_%Y%m%d_%H%M%S.log"))
    utils.configurar_logging(ruta_log_file)

    ruta_excel = os.path.join(ruta_mes,"Solicitud.xlsx")
    if not os.path.isfile(ruta_excel):
        utils.print_error(f"No se encontró archivo Excel en {ruta_excel}")
        return

    ruta_bd_docentes = os.path.join(config.RAIZ, "BD-DOCENTES.xlsx")
    if os.path.isfile(ruta_bd_docentes):
        stats_doc = docente_repository.sync_docentes_from_excel(ruta_bd_docentes)
        utils.print_info(
            f"Docentes sincronizados desde BD-DOCENTES.xlsx: +{stats_doc['inserted']} nuevos, "
            f"{stats_doc['updated']} actualizados."
        )
        logging.info("sync_docentes stats=%s", stats_doc)
    else:
        utils.print_warning("No se encontró BD-DOCENTES.xlsx en la raíz; se mantiene catálogo actual de docentes.")

    continuar = utils.mostrar_contexto_ejecucion(
        "🗂️ Contexto de ejecución",
        [
            ("Raíz", RAIZ),
            ("Período", f"{mes} {año}"),
            ("Carpeta mes", ruta_mes),
            ("Excel", ruta_excel),
            ("BD docentes", ruta_bd_docentes),
            ("Logs", ruta_log_file),
        ],
        preview_items=["Se validará recepción PDF/XML y se actualizará la hoja seleccionada."],
        confirm_message="¿Continuar con la validación de recepción? (s/n)",
    )
    if not continuar:
        utils.print_warning("Proceso cancelado por el usuario.")
        return

    wb = load_workbook(ruta_excel, read_only=True)
    hojas = wb.sheetnames
    wb.close()
    hoja = utils.choose_excel_sheet(
        hojas,
        sheet=getattr(args, "sheet", None),
        prompt_message="Seleccione la hoja del Excel para validar:",
    )
    df = pd.read_excel(ruta_excel, sheet_name=hoja, engine='openpyxl')

    # Validación canónica del esquema (Fase 3)
    canonical_errors, canonical_warnings = schema_validator.validate_for_stage(
        df, "stage3_validacion_recepcion"
    )
    for w in canonical_warnings:
        logging.warning(f"[script3] WARN {w}")
        utils.print_warning(f"[schema] {w}")
    for e in canonical_errors:
        logging.error(f"[script3] ERROR {e}")
        utils.print_error(f"[schema] {e}")
    if canonical_errors and getattr(args, "strict", False):
        utils.print_error("Validación estricta activada y se detectaron errores de esquema. Abortando.")
        return

    logging.info(f"Iniciando proceso para {año}/{mes}, hoja: {hoja}")
    df_actualizado = procesar_filas(df, ruta_mes)

    periodo_id = None
    if mes_num > 0:
        periodo_id = file_repository.get_or_create_periodo(
            anio=int(año),
            mes_num=mes_num,
            mes_nombre=mes,
        )
    for idx, fila in df_actualizado.iterrows():
        boleta_repository.upsert_boleta_recepcion(
            periodo_id=periodo_id,
            boleta_key=build_boleta_key(fila.to_dict()),
            emplid=str(fila.get("EMPLID", "")).strip() if "EMPLID" in df_actualizado.columns else None,
            rut_sin_dv=str(fila.get("RUT_SIN_DV", "")).strip() if "RUT_SIN_DV" in df_actualizado.columns else None,
            rut_razon=str(fila.get("RUT RAZON", "")).strip() if "RUT RAZON" in df_actualizado.columns else None,
            estado_recepcion=str(fila.get("Estado_Recepcion", "")).strip() or None,
            observaciones_recepcion=str(fila.get("Observaciones", "")).strip() or None,
            glosa=str(fila.get("GLOSA", "")).strip() if "GLOSA" in df_actualizado.columns else None,
            monto_bruto=fila.get("CUS_TOT_HON", None),
            archivo_xml=str(fila.get("archivo_xml", "")).strip() if "archivo_xml" in df_actualizado.columns else None,
        )

    reporte = generar_reporte_texto(df_actualizado)
    ruta_reporte_dir = os.path.join(ruta_mes, "reporte_avance")
    os.makedirs(ruta_reporte_dir, exist_ok=True)
    ruta_reporte = os.path.join(
        ruta_reporte_dir,
        f"reporte_revision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    )
    with open(ruta_reporte, "w", encoding="utf-8") as f:
        f.write(reporte)
    utils.print_success(f"Reporte generado en: {ruta_reporte}")

    guardar_excel(df_actualizado, ruta_excel, hoja)

    logging.info("Proceso completado.")
    utils.print_success("🎯 Validación finalizada correctamente. Revisa los logs para detalles.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Validación de recepción PDF/XML")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Aborta si la Solicitud.xlsx no cumple el esquema canónico.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Nombre de la hoja del Excel (ej. Solicitud).",
    )
    utils.register_non_interactive_cli(parser)
    utils.register_period_args(parser)
    args = parser.parse_args()
    main(args)
