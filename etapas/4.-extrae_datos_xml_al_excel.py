#!/usr/bin/env python3
import _sys_path  # noqa: E402
import os
import re
import argparse
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import logging
from datetime import datetime
from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn
from rich.panel import Panel
import utils
import schema_validator
from db import file_repository, xml_repository
from db.key_builder import build_boleta_key

console = utils.console

import config
RAIZ = config.RAIZ
PREFIJO = config.PREFIJO

EXTRAER_CAMPOS = [
    'rutEmisor', 'dvEmisor', 'rutReceptor', 'dvReceptor', 'nombreReceptor',
    'totalHonorarios', 'liquidoHonorarios', 'impuestoHonorarios', 'descripcionLinea',
    'fechaBoleta', 'numeroBoleta', 'porcentajeImpuesto'
]

COLUMNAS_OBJETIVO = [
    'rutEmisorCompleto_XML',
    'rutReceptorCompleto_XML',
    'nombreReceptor_XML',
    'porcentajeImpuesto_XML',
    'totalHonorarios_XML',
    'liquidoHonorarios_XML',
    'impuestoHonorarios_XML',
    'descripcionLinea_XML',
    'fechaBoleta_XML',
    'numeroBoleta_XML',
    'Archivo_XML_Usado',
    'Observaciones_XML'
]

def crear_columnas_si_no_existen(df):
    for col in COLUMNAS_OBJETIVO:
        if col not in df.columns:
            df[col] = ''

def limpiar_columnas_objetivo(df):
    for col in COLUMNAS_OBJETIVO:
        if col in df.columns:
            df.drop(columns=[col], inplace=True)

def extraer_datos_completos_xml(ruta_xml):
    try:
        tree = ET.parse(ruta_xml)
        root = tree.getroot()
        datos = {}
        for campo in EXTRAER_CAMPOS:
            elem = utils.find_element_ignore_ns(root, campo)
            datos[campo] = elem.text.strip() if elem is not None and elem.text else ''
        return datos
    except (ET.ParseError, OSError, ValueError, TypeError) as e:
        return {'error': f"Error al leer XML: {e}"}

def seleccionar_opcion(lista, mensaje, icono=""):
    # Delegar a utils para mantener una única implementación compartida
    return utils.seleccionar_opcion(lista, mensaje, icono)

def main(args=None):
    if args is None:
        args = argparse.Namespace(strict=False, yes=False, year=None, month=None)
    utils.apply_non_interactive_from_args(args)
    utils.print_header("📂 EXTRACCIÓN DE DATOS XML AL EXCEL")

    try:
        año, mes = utils.resolve_año_mes(RAIZ, getattr(args, "year", None), getattr(args, "month", None))
    except ValueError as e:
        utils.print_error(str(e))
        return
    ruta_mes = os.path.join(RAIZ, año, mes)
    mes_num = config.MESES_ES.index(mes) + 1 if mes in config.MESES_ES else 0
    periodo_id = None
    if mes_num > 0:
        periodo_id = file_repository.get_or_create_periodo(
            anio=int(año),
            mes_num=mes_num,
            mes_nombre=mes,
        )

    ruta_logs = os.path.join(ruta_mes, "logs_extraccion_xml_excel")
    os.makedirs(ruta_logs, exist_ok=True)
    ruta_log_file = os.path.join(ruta_logs, datetime.now().strftime("extraccion_%Y%m%d_%H%M%S.log"))
    utils.configurar_logging(ruta_log_file)
    
    ruta_excel = os.path.join(ruta_mes, "Solicitud.xlsx")
    if not os.path.isfile(ruta_excel):
        utils.print_error(f"No se encontró archivo Excel en {ruta_excel}")
        return

    continuar = utils.mostrar_contexto_ejecucion(
        "🗂️ Contexto de ejecución",
        [
            ("Raíz", RAIZ),
            ("Período", f"{mes} {año}"),
            ("Carpeta mes", ruta_mes),
            ("Excel", ruta_excel),
            ("Logs", ruta_log_file),
        ],
        preview_items=["Se extraerán datos XML y se sobrescribirá la hoja seleccionada."],
        confirm_message="¿Continuar con la extracción de XML al Excel? (s/n)",
    )
    if not continuar:
        utils.print_warning("Proceso cancelado por el usuario.")
        return

    wb = load_workbook(ruta_excel, read_only=True)
    hojas = wb.sheetnames
    wb.close()
    hoja = utils.choose_excel_sheet(
        hojas,
        sheet=getattr(args, "sheet", None),
        prompt_message="Seleccione la hoja del Excel para agregar datos:",
    )
    df = pd.read_excel(ruta_excel, sheet_name=hoja, engine='openpyxl')

    # Validación canónica del esquema (Fase 3)
    canonical_errors, canonical_warnings = schema_validator.validate_for_stage(
        df, "stage4_extraccion_xml"
    )
    for w in canonical_warnings:
        logging.warning(f"[script4] WARN {w}")
        utils.print_warning(f"[schema] {w}")
    for e in canonical_errors:
        logging.error(f"[script4] ERROR {e}")
        utils.print_error(f"[schema] {e}")
    if canonical_errors and getattr(args, "strict", False):
        utils.print_error("Validación estricta activada y se detectaron errores de esquema. Abortando.")
        return

    crear_columnas_si_no_existen(df)

    # Convertir columnas a tipo string
    for col in COLUMNAS_OBJETIVO:
        df[col] = df[col].astype(str)

    hay_datos_ok = df['Observaciones_XML'].astype(str).str.strip().str.lower().eq('datos extraídos ok').any()
    sobrescribir_ok = False
    if hay_datos_ok:
        if utils.is_non_interactive():
            sobrescribir_ok = False
            utils.print_info("Modo no interactivo: no se sobrescriben filas con 'Datos extraídos OK'.")
        else:
            utils.print_warning("¿Desea sobrescribir filas que ya tienen 'Datos extraídos OK'?")
            respuesta = utils.prompt_required("Ingrese S para sí, N para no").strip().lower()
            sobrescribir_ok = respuesta == 's'
        if sobrescribir_ok:
            limpiar_columnas_objetivo(df)
            crear_columnas_si_no_existen(df)

    total = len(df)
    exitos = 0
    errores = 0

    utils.print_progress_status(f"Iniciando extracción de datos XML para {total} filas...")

    with Progress(
        "[progress.description]{task.description}",
        BarColumn(),
        "[progress.percentage]{task.percentage:>3.1f}%",
        "•",
        TimeElapsedColumn(),
        console=console,
        transient=True,
    ) as progress:
        task = progress.add_task("[green]Procesando filas...", total=total)
        for idx, fila in df.iterrows():
            archivo_xml = str(fila.get('archivo_xml', '')).strip()
            observacion_actual = str(fila.get('Observaciones_XML', '')).strip()

            if observacion_actual.lower() == 'datos extraídos ok' and not sobrescribir_ok:
                progress.advance(task)
                continue

            if not archivo_xml:
                if observacion_actual != 'Sin archivo XML relacionado':
                    df.at[idx, 'Observaciones_XML'] = 'Sin archivo XML relacionado'
                progress.advance(task)
                continue

            ruta_archivo_xml = os.path.join(ruta_mes, archivo_xml)
            if not os.path.isfile(ruta_archivo_xml):
                df.at[idx, 'Observaciones_XML'] = 'Archivo XML no encontrado'
                errores += 1
                progress.advance(task)
                continue

            datos = extraer_datos_completos_xml(ruta_archivo_xml)
            if 'error' in datos:
                df.at[idx, 'Observaciones_XML'] = datos['error']
                errores += 1
                progress.advance(task)
                continue

            # Guardar datos básicos
            df.at[idx, 'rutEmisorCompleto_XML'] = datos.get('rutEmisor', '') + datos.get('dvEmisor', '')
            rut_receptor_xml = datos.get('rutReceptor', '') + datos.get('dvReceptor', '')
            df.at[idx, 'rutReceptorCompleto_XML'] = rut_receptor_xml
            df.at[idx, 'nombreReceptor_XML'] = datos.get('nombreReceptor', '')
            df.at[idx, 'porcentajeImpuesto_XML'] = datos.get('porcentajeImpuesto', '')
            df.at[idx, 'totalHonorarios_XML'] = datos.get('totalHonorarios', '')
            df.at[idx, 'liquidoHonorarios_XML'] = datos.get('liquidoHonorarios', '')
            df.at[idx, 'impuestoHonorarios_XML'] = datos.get('impuestoHonorarios', '')
            df.at[idx, 'descripcionLinea_XML'] = datos.get('descripcionLinea', '')
            df.at[idx, 'fechaBoleta_XML'] = datos.get('fechaBoleta', '')
            df.at[idx, 'numeroBoleta_XML'] = datos.get('numeroBoleta', '')
            df.at[idx, 'Archivo_XML_Usado'] = archivo_xml

            # Nuevas validaciones
            observaciones = []
            # Validar monto
            try:
                monto_excel = float(fila.get('CUS_TOT_HON', 0))
                monto_xml = float(datos.get('totalHonorarios', 0))
                if abs(monto_excel - monto_xml) > 0.01:
                    observaciones.append(f"Monto Excel ({monto_excel}) distinto a XML ({monto_xml})")
            except (TypeError, ValueError):
                observaciones.append("Error conversión monto")

            # Validar RUT receptor con guion antes del DV
            rut_receptor_xml_con_guion = f"{datos.get('rutReceptor','')}-{datos.get('dvReceptor','')}"
            rut_receptor_excel = str(fila.get('RUT RAZON', '')).strip()
            if rut_receptor_xml_con_guion != rut_receptor_excel:
                observaciones.append(f"RUT receptor Excel ({rut_receptor_excel}) distinto a XML ({rut_receptor_xml_con_guion})")

            if observaciones:
                df.at[idx, 'Observaciones_XML'] = '; '.join(observaciones)
                errores += 1
            else:
                df.at[idx, 'Observaciones_XML'] = 'Datos extraídos OK'
                exitos += 1

            estado_rx = str(fila.get("Estado_Recepcion", "")).strip().upper()
            if estado_rx in {"RECIBIDO", "RECIBIDO CON ERROR"} and archivo_xml:
                xml_repository.upsert_boleta_xml_data(
                    periodo_id=periodo_id,
                    boleta_key=build_boleta_key(fila.to_dict(), row_index=idx),
                    emplid=str(fila.get("EMPLID", "")).strip() if "EMPLID" in df.columns else None,
                    rut_sin_dv=str(fila.get("RUT_SIN_DV", "")).strip() if "RUT_SIN_DV" in df.columns else None,
                    datos=datos,
                    observaciones_xml=str(df.at[idx, 'Observaciones_XML']).strip() or None,
                )

            progress.advance(task)

    try:
        hojas = pd.read_excel(ruta_excel, sheet_name=None, engine='openpyxl')
        hojas[hoja] = df

        def _writer(tmp_path):
            with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='w') as writer:
                for nombre_hoja, df_hoja in hojas.items():
                    df_hoja.to_excel(writer, index=False, sheet_name=nombre_hoja)

        utils.atomic_excel_write(ruta_excel, _writer)
    except PermissionError as e:
        utils.print_error(f"[BH-EXCEL_LOCKED] El archivo {ruta_excel} está abierto o bloqueado: {e}")
        return
    except (OSError, IOError, ValueError, KeyError) as e:
        utils.print_error(f"[BH-EXCEL_WRITE] No se pudo guardar el Excel: {e}")
        logging.exception("Fallo al escribir Excel")
        return

    resumen = (
        f"✅ Filas procesadas: {total}\n"
        f"✔️ Extracciones exitosas: {exitos}\n"
        f"⚠️ Errores encontrados: {errores}\n"
        f"📄 Archivo Excel sobrescrito correctamente."
    )
    console.print(Panel(resumen, title="Resumen de extracción", style="bold green"))
    logging.info("Extracción completada exitosamente.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extracción de datos XML al Excel")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Aborta si la Solicitud.xlsx no cumple el esquema canónico.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Nombre de la hoja del Excel (ej. Solicitud).",
    )
    utils.register_non_interactive_cli(parser)
    utils.register_period_args(parser)
    args = parser.parse_args()
    main(args)
